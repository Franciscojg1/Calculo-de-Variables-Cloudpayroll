#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
SCRIPT DE CÁLCULO DE VARIABLES DE LIQUIDACIÓN - VERSIÓN MEJORADA 2.0

Características principales:
- Implementa todas las reglas del documento REGLAS.docx
- Sistema de logging detallado para debugging
- Validaciones exhaustivas de datos de entrada
- Generación de reportes de procesamiento
- Manejo robusto de errores
- Documentación clara de cada función
"""

import json
import unicodedata
import math
import logging
import os
import re
from datetime import datetime
import traceback
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Any, Dict, List, Optional, Set, Tuple
from typing import Callable, Any, Dict, List, Optional, Tuple
from collections import defaultdict

logger = logging.getLogger('json_a_excel')

# Desactivada temporalmente: mantener la lógica para una eventual reactivación.
VARIABLE_1151_HABILITADA = False

def json_a_excel_streamlit(ruta_json: str, nombre_excel: str = "variables_calculadas.xlsx", logger_callback=None) -> Optional[str]:
    """
    Procesa un archivo JSON normalizado (legajos) y genera un Excel con variables calculadas.
    Retorna el path del Excel generado, o None si hubo error crítico.
    """
    try:
        # 2. Leer el archivo JSON
        if not os.path.exists(ruta_json):
            logger.error(f"No se encontró el archivo: {ruta_json}")
            if logger_callback: logger_callback(f"No se encontró el archivo: {ruta_json}")
            return None

        with open(ruta_json, "r", encoding="utf-8") as f:
            data = json.load(f)

        if "legajos" not in data:
            logger.error("El JSON no contiene la clave 'legajos'")
            if logger_callback: logger_callback("El JSON no contiene la clave 'legajos'")
            return None
    except Exception as e:
            logger.error(f"Ocurrió un error crítico procesando el JSON: {e}", exc_info=True)
    if logger_callback:
        logger_callback(f"Ocurrió un error crítico procesando el JSON: {e}")
    return None

# --- Códigos de Color ANSI (deshabilitados para Streamlit) ---
COLOR_RESET = ""
COLOR_BLACK = ""
COLOR_RED = ""
COLOR_GREEN = ""
COLOR_YELLOW = ""
COLOR_BLUE = ""
COLOR_MAGENTA = ""
COLOR_CYAN = ""
COLOR_WHITE = ""
COLOR_BOLD = ""
COLOR_UNDERLINE = ""


def normalizar_texto(texto: Any) -> str:
    """
    Normaliza un texto para comparaciones robustas.
    """
    if texto is None:
        return ""

    if not isinstance(texto, str):
        texto = str(texto) if texto else ""

    try:
        texto_procesado = texto.lower()

        reemplazos_directos = {
            'ñ': 'n',
            'ç': 'c',
        }
        for char, reemplazo in reemplazos_directos.items():
            texto_procesado = texto_procesado.replace(char, reemplazo)

        texto_normalizado_unicode = unicodedata.normalize('NFKD', texto_procesado)

        texto_sin_diacriticos = ''.join(
            c for c in texto_normalizado_unicode
            if not unicodedata.combining(c)
        )

        texto_filtrado = re.sub(r'[^a-z0-9\s]', ' ', texto_sin_diacriticos)

        texto_limpio = re.sub(r'\s+', ' ', texto_filtrado).strip()

        return texto_limpio

    except Exception as e:
        # Puedes decidir si quieres mantener solo un logger.error aquí para casos de falla
        logger.error(f"Error crítico al normalizar texto: '{texto}'. Error: {str(e)}", exc_info=True)
        return str(texto).lower().strip()

def print_header():
    """Imprime el encabezado del programa"""
    header = """
    ============================================================
    SCRIPT DE CÁLCULO DE VARIABLES DE LIQUIDACIÓN - VERSIÓN 2.0
    Sistema automatizado para cálculo preciso de variables según:
    - Documento REGLAS.docx (Única fuente de verdad)
    - Estructura JSON normalizada
    ============================================================
    """
    print(header)
    logger.info("Inicializando sistema de cálculo de variables")

# ---
# ======================
# CONSTANTES NORMALIZADAS
# ======================

# Puestos especiales con nombres normalizados
PUESTOS_ESPECIALES: Dict[str, str] = {
    'TELEFONISTA': normalizar_texto('TELEFONISTA'),
    'RECEP_LAB': normalizar_texto('RECEPCIONISTA DE LABORATORIO'),
    'TEC_CARDIO': normalizar_texto('TECNICO EN PRACTICAS CARDIOLOGICAS'),
    'OP_LOGISTICA': normalizar_texto('OPERARIO DE LOGISTICA'),
    'MEDICO': normalizar_texto('MEDICO'),
    'MEDICA': normalizar_texto('MEDICA'),
    'MEDICO/A': normalizar_texto('MEDICO/A'),
    'ODONTOLOGO': normalizar_texto('ODONTOLOGO/A'),
    'ODONTOLOGO/A FELLOW': normalizar_texto('ODONTOLOGO/A FELLOW')
}

valores_profesionales_para_comparacion = (
            normalizar_texto('MEDICO'),
            normalizar_texto('MEDICA'),
            normalizar_texto('MEDICO/A'), # Normalizado al vuelo
            normalizar_texto('ODONTOLOGO/A'), # Normalizado al vuelo
            normalizar_texto('ODONTOLOGO/A FELLOW') # Normalizado al vuelo
        )

# Constantes de PISOS HORARIOS (claves en minúsculas)
PISOS_HORARIOS: Dict[str, float] = {
    normalizar_texto('GENERAL'): 36.0,
    normalizar_texto('LABORATORIO'): 27.0,
    normalizar_texto('IMAGENES'): 18.0
}

# Sector de Laboratorio excluido (en minúsculas)
SECTOR_EXCLUIDO_LABORATORIO = normalizar_texto("Laboratorio")

# Conjuntos de sectores imágenes (valores en minúsculas)
SECTORES_IMAGENES: Set[str] = {
    normalizar_texto("MAMOGRAFIA"),
    normalizar_texto("IMAGENES DMF"),
    normalizar_texto("TOMOGRAFIA COMPUTADA"),
    normalizar_texto("DENSITOMETRIA"),
    normalizar_texto("MEDICINA NUCLEAR"),
    normalizar_texto("PET/CT"),
    normalizar_texto("RADIOLOGIA"),
    normalizar_texto("RESONANCIA MAGNETICA"),
    normalizar_texto("IMAGENES") # Aseguramos que 'IMAGENES' esté si se usa como clave general
}

# Sectores con reglas especiales (valores en minúsculas)
SECTORES_ESPECIALES: Dict[str, List[str]] = {
    'HORAS_200': [normalizar_texto("CUAT")],
    'HORAS_156': [
        normalizar_texto("LABORATORIO"),
        normalizar_texto("MAMOGRAFIA"),
        normalizar_texto("IMAGENES DMF"),
        normalizar_texto("TOMOGRAFIA COMPUTADA"),
        normalizar_texto("DENSITOMETRIA"),
        normalizar_texto("MEDICINA NUCLEAR"),
        normalizar_texto("PET/CT"),
        normalizar_texto("RADIOLOGIA"),
        normalizar_texto("RESONANCIA MAGNETICA")
    ],
    'MEDICOS': [
        normalizar_texto("ECOGRAFIA"),
        normalizar_texto("MAMOGRAFIA")
    ]
}

# Términos especiales en horarios (valores en minúsculas)
TERMINOS_ESPECIALES: Set[str] = {
    normalizar_texto("SADOFE"),
    normalizar_texto("DOFE"),
    normalizar_texto("SADO"),
    normalizar_texto("SAFE")
}

# Sedes que no liquidan plus guardia (valores en minúsculas)
SEDES_NO_LIQUIDA_PLUS: Set[str] = {
    normalizar_texto("CLINICA BAZTERRICA"),
    normalizar_texto("CLINICA DEL SOL"),
    normalizar_texto("CONSULTORIOS BAZTERRICA"),
    normalizar_texto("PATERNAL"),
    normalizar_texto("C DEL SOL"),
    normalizar_texto("CDS"),
    normalizar_texto("C. DEL SOL")
}
# Sedes que están permitidas para considerar a alguien full guardia (valores en minúsculas)
sedes_permitidas = {
    normalizar_texto('clínica del sol'),
    normalizar_texto('c. del sol'),
    normalizar_texto('cds'),
    normalizar_texto('san miguel'),
    normalizar_texto('sm'),
    normalizar_texto('bazterrica'),
    normalizar_texto('cons. ext. cl. bazterrica'),
    normalizar_texto('clinica bazterrica'),
    normalizar_texto('Cons. Ext. Cl. Bazterrica'),
    normalizar_texto('Santa Isabel'),
    normalizar_texto('Clinica Santa Isabel'),
    normalizar_texto('Paternal')
}
# Constantes específicas para es_medico_productividad (valores en minúsculas)
SECTORES_MEDICOS: Set[str] = {
    normalizar_texto("ECOGRAFIA"),
    normalizar_texto("MAMOGRAFIA")
}

DIAS_ESPECIALES = {0, 1, 2}  # Lunes, Martes, Miércoles

# ======================
# REGLAS ESPECIALES - CLASES DE CONFIGURACIÓN
# ======================

class ConfigArt19:
    """Configuraciones para cálculo de Artículo 19"""
    PUESTOS_VALIDOS: Set[str] = {
        normalizar_texto("TECNICO DE LABORATORIO"),
        normalizar_texto("EXTRACCIONISTA"),
        normalizar_texto("BIOQUIMICO"),
        normalizar_texto("AUXILIAR TÉCNICO")
    }
    SECTOR_VALIDO: str = normalizar_texto("LABORATORIO")
    CATEGORIA_PREFIX: str = 'dc_' # Esta se compara con .lower(), así que el prefijo es lowercase
    HORAS_MIN: float = 36.0
    HORAS_MAX: float = 48.0
    PORCENTAJE_MAX: float = 33 # Variable antes CONSTANTES['PORCENTAJE_MAX_ART19']

class ConfigExtensionHoraria:
    """Configuraciones para extensión horaria (Variable 992)"""
    PUESTOS_VALIDOS: Set[str] = {
        normalizar_texto("TECNICO"),
        normalizar_texto("TECNICO PIVOT")
    }
    HORAS_MINIMAS: float = 24.0
    ID_LEGAJO_EXCLUIDO_MIN: int = 4000
    ID_LEGAJO_EXCLUIDO_MAX: int = 4999

class ConfigBioimagenes:
    """Configuraciones para adicional de bioimágenes (Variable 10000)"""
    PUESTOS_VALIDOS: Set[str] = {
        normalizar_texto("TECNICO"),
        normalizar_texto("TECNICO DE REPROCESO"),
        normalizar_texto("TECNICO PIVOT")
    }
    TERMINOS_ADICIONALES: Set[str] = {
        normalizar_texto("LIC. EN BIOIMAGENES"),
        normalizar_texto("BIOIMAGENES"),
        normalizar_texto("LICENCIADO EN BIOIMAGENES"),
        normalizar_texto("PRESENTÓ TÍTULO"),
        normalizar_texto("TÍTULO")
    }

class ConfigAdicionalPivot:
    """Configuraciones para adicional pivot (Variables 1145 y 1144)."""
    PUESTO_VALIDO: str = normalizar_texto("TECNICO PIVOT")
    SECTOR_RESONANCIA: str = normalizar_texto("RESONANCIA MAGNETICA")
    SECTORES_VARIABLE_1144: Set[str] = {
        normalizar_texto("TOMOGRAFIA COMPUTADA"),
        normalizar_texto("IMAGENES DMF"),
        normalizar_texto("CHEQUEOS Y CARDIOLOGIA"),
        normalizar_texto("MAMOGRAFIA"),
        normalizar_texto("DENSITOMETRIA"),
        normalizar_texto("RADIOLOGIA")
    }
    VARIABLE_RESONANCIA: int = 1145
    VALOR_RESONANCIA: int = 40
    VARIABLE_GENERAL: int = 1144
    VALOR_GENERAL: int = 20

# Variables utilizadas en calcular_porcentaje_art19
CATEGORIA_ART19_PREFIX: str = ConfigArt19.CATEGORIA_PREFIX
PUESTOS_ART19: Set[str] = ConfigArt19.PUESTOS_VALIDOS
SECTOR_ART19: str = ConfigArt19.SECTOR_VALIDO
HORAS_MIN_ART19: float = ConfigArt19.HORAS_MIN
HORAS_MAX_ART19: float = ConfigArt19.HORAS_MAX
CONSTANTES: Dict[str, float] = {'PORCENTAJE_MAX_ART19': ConfigArt19.PORCENTAJE_MAX}
HORAS_BASE_CALCULO_ART19: float = 48.0 # Asumiendo 48 horas como base para el cálculo proporcional

TERMINOS_CESION_RAW = [
    "Cesión",
    "CECION" 
]

# Y luego normalizar la lista para crear el set final
# Esto se hace una sola vez cuando el script se carga
TERMINOS_CESION = {normalizar_texto(term) for term in TERMINOS_CESION_RAW}

# ======================
# CATÁLOGO COMPLETO DE VARIABLES
# ======================

CATALOGO_VARIABLES = {
    1: "Sueldo Bruto Pactado",
    4: "Horas Mensuales",
    239: "Horas Semanales",
    426: "Cajero/Seguro",
    992: "Extensión Horaria",
    1131: "Días Especiales Mensuales",
    1137: "Lavado de Uniforme",
    1145: "Adicional Voluntario Empr. Reso. (Pivot)",
    1151: "Adicional Resonancia Magnética",
    1157: "Horas Nocturnas Mensuales",
    1167: "Jornada Reducida (%)",
    1242: "Días Mensuales",
    1251: "Médico Productividad (Flag 1)",
    1252: "Médico Productividad (Flag 2)",
    1416: "Jornada Art. 19",
    1498: "Adicional Nocturno",
    1599: "Porcentaje Art. 19",
    1673: "Proporción Lavado",
    1740: "Médico Productividad (Principal)",
    1144: "Adicional Voluntario Emp. (Pivot)",
    2000: "Personal de Guardia",
    2006: "Fecha Fin de Contrato",
    2281: "No Liquida Plus Guardia",
    7000: "Info: Es Cesión",
    8000: "Info: Revisar Intangibilidad",
    9000: "Info: Adicional Voluntario",
    10000: "Info: Licenciado Bioimágenes",
    11000: "Info: PPR - Revisar Archivo",
    12000: "Info: Falta Sueldo Bruto PFC",
    13000: "Info: Guardias de Capacitación"
}

# ======================
# FUNCIONES DE LOGGING ESTANDARIZADAS
# ======================

def log_variable_calculada(id_legajo: Any, cod_variable: int, valor: Any, razon: str = "") -> None:
    """
    Log estandarizado para variables CALCULADAS.
    
    Args:
        id_legajo: ID del legajo
        cod_variable: Código de la variable
        valor: Valor calculado
        razon: Razón opcional del cálculo
    """
    nombre_var = CATALOGO_VARIABLES.get(cod_variable, f"V{cod_variable}")
    razon_texto = f" - {razon}" if razon else ""
    
    mensaje = f"V{cod_variable} ({nombre_var}): ✓ CALCULADA = {valor}{razon_texto}"
    logger.info(f"Legajo {id_legajo}: {mensaje}")

def log_variable_no_calculada(id_legajo: Any, cod_variable: int, razon: str) -> None:
    """
    Log estandarizado para variables NO CALCULADAS.
    
    Args:
        id_legajo: ID del legajo
        cod_variable: Código de la variable
        razon: Razón por la que no se calculó
    """
    nombre_var = CATALOGO_VARIABLES.get(cod_variable, f"V{cod_variable}")
    
    mensaje = f"V{cod_variable} ({nombre_var}): ✗ NO CALCULADA - {razon}"
    logger.debug(f"Legajo {id_legajo}: {mensaje}")

def log_variable_evaluando(id_legajo: Any, cod_variable: int) -> None:
    """
    Log para indicar que se está evaluando una variable.
    
    Args:
        id_legajo: ID del legajo
        cod_variable: Código de la variable
    """
    # Comentado para reducir ruido en logs
    # nombre_var = CATALOGO_VARIABLES.get(cod_variable, f"V{cod_variable}")
    # logger.debug(f"Legajo {id_legajo}: Evaluando V{cod_variable} ({nombre_var})...")
    pass

def log_resumen_variables(id_legajo: Any, variables: List[Tuple[int, Any]]) -> None:
    """
    Log de resumen final con todas las variables calculadas.
    
    Args:
        id_legajo: ID del legajo
        variables: Lista de tuplas (codigo_variable, valor)
    """
    logger.info(f"\n{'='*80}")
    logger.info(f"RESUMEN DE VARIABLES CALCULADAS - Legajo {id_legajo}")
    logger.info(f"{'='*80}")
    
    if not variables:
        logger.info(f"No se calcularon variables para este legajo")
        return
    
    # Ordenar por código de variable
    variables_ordenadas = sorted(variables, key=lambda x: x[0])
    
    for cod_var, valor in variables_ordenadas:
        nombre_var = CATALOGO_VARIABLES.get(cod_var, f"V{cod_var}")
        mensaje = f"  ✓ V{cod_var:4d} ({nombre_var:40s}): {valor}"
        logger.info(mensaje)
    
    logger.info(f"{'='*80}")
    logger.info(f"Total variables calculadas: {len(variables)}\n")

# ==============================
# FUNCIONES PRINCIPALES
# ==============================

def procesar_archivo_json(
    ruta_archivo: str,
    modo_resumen: str = "mixto",  # "mixto" | "normalizado" | "crudo"
) -> Tuple[Optional[List[Tuple[int, int, Any]]], Dict[str, Any], Dict[Any, Any]]:
    """
    Procesa el archivo JSON y genera:
      - resultados: Lista de tuplas (id_legajo, codigo_variable, valor) o None
      - stats: métricas del procesamiento
      - resumen_horarios: dict {id_legajo: info_enriquecida}

    modo_resumen:
      - "mixto": prioriza campos normalizados y hace fallback al crudo si faltan (recomendado)
      - "normalizado": siempre usa los campos normalizados
      - "crudo": siempre usa los campos crudos (horario_resumen se desactiva)
    """
    logger = logging.getLogger('json_a_excel')

    # Helpers internos para selección de valores
    def _is_missing(v):
        if v is None:
            return True
        if isinstance(v, str) and v.strip() == "":
            return True
        if isinstance(v, (list, dict)) and len(v) == 0:
            return True
        return False

    def pick(norm, raw):
        if modo_resumen == "normalizado":
            return norm
        if modo_resumen == "crudo":
            return raw
        # mixto (default)
        return raw if _is_missing(norm) else norm

    # Inicialización de estadísticas
    stats: Dict[str, Any] = {
        'total_legajos': 0,
        'legajos_procesados': 0,
        'legajos_con_error': 0,
        'variables_calculadas': 0,
        'errores_por_tipo': defaultdict(int),
    }

    resumen_horarios: Dict[Any, Any] = {}

    try:
        logger.info(f"📂 Cargando archivo JSON: {ruta_archivo}")
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if 'legajos' not in data:
            error_msg = "El archivo JSON no contiene la clave 'legajos'"
            logger.error(error_msg)
            return None, stats, resumen_horarios

        stats['total_legajos'] = len(data['legajos'])
        resultados: List[Tuple[int, int, Any]] = []
        logger.info(f"🔍 Iniciando procesamiento de {stats['total_legajos']} legajos")

        for i, legajo in enumerate(data['legajos'], 1):
            # ----------- Armado del resumen enriquecido -----------
            crudo = legajo.get('crudo_min', {}) or {}
            dp = legajo.get('datos_personales', {}) or {}
            contr = legajo.get('contratacion', {}) or {}
            fechas = contr.get('fechas', {}) or {}
            remu = legajo.get('remuneracion', {}) or {}
            hor = legajo.get('horario', {}) or {}

            legajo_id = (
                legajo.get('id_legajo')
                or crudo.get('Legajo')
                or legajo.get('legajo')
                or legajo.get('id')
                or 'DESCONOCIDO'
            )

            # sector puede venir como dict en datos_personales
            sector_dict = dp.get('sector') if isinstance(dp.get('sector'), dict) else {}
            sector_principal_norm = sector_dict.get('principal') if sector_dict else None
            sector_sub_norm = sector_dict.get('subsector') if sector_dict else None

            resumen_horarios[legajo_id] = {
                'nombre_completo': pick(dp.get('nombre'), crudo.get('Nombre completo')),
                'sector': pick(sector_principal_norm, crudo.get('Sector')),
                'subsector': pick(sector_sub_norm, crudo.get('Subsector')),
                'puesto': pick(dp.get('puesto'), crudo.get('Puesto')),
                'sede': pick(dp.get('sede'), crudo.get('Sede')),
                'categoria': pick(contr.get('categoria'), crudo.get('Categoría')),
                'modalidad': pick(contr.get('tipo'), crudo.get('Modalidad contratación')),
                'fecha_ingreso': pick(fechas.get('ingreso'), crudo.get('Fecha ingreso')),
                'fecha_fin': pick(fechas.get('fin'), crudo.get('Fecha de fin')),
                'sueldo_bruto_pactado': pick(remu.get('sueldo_base'), crudo.get('Sueldo bruto pactado')),
                'adicionales': pick(remu.get('adicionables'), crudo.get('Adicionales')),
                # Horario: texto crudo (o texto_original si está disponible), y resumen solo si no es modo "crudo"
                'horario_texto': (
                    crudo.get('Horario completo') if modo_resumen == "crudo"
                    else (hor.get('texto_original') or crudo.get('Horario completo'))
                ),
                'horario_resumen': None if modo_resumen == "crudo" else hor.get('resumen'),
            }
            # ----------- Fin resumen enriquecido -----------

            try:
                logger.debug(f"Procesando legajo {i}/{stats['total_legajos']} (ID: {legajo_id})")

                if not validar_estructura_legajo(legajo):
                    stats['legajos_con_error'] += 1
                    stats['errores_por_tipo']['estructura_invalida'] += 1
                    logger.warning(f"Estructura inválida en legajo {legajo_id}")
                    continue

                variables_legajo = calcular_variables(legajo)
                if not variables_legajo:
                    logger.debug(f"Legajo {legajo_id} no generó variables calculadas")
                    continue

                for var_codigo, var_valor in variables_legajo:
                    resultados.append((legajo_id, var_codigo, var_valor))

                stats['legajos_procesados'] += 1
                stats['variables_calculadas'] += len(variables_legajo)

                if i % 10 == 0:
                    logger.info(
                        f"📊 Progreso: {i}/{stats['total_legajos']} | "
                        f"Éxitos: {stats['legajos_procesados']} | Errores: {stats['legajos_con_error']}"
                    )

            except Exception as e:
                stats['legajos_con_error'] += 1
                stats['errores_por_tipo'][type(e).__name__] += 1
                logger.error(f"⚠ Error procesando legajo {legajo_id}: {str(e)}")
                try:
                    logger.debug(f"Datos legajo problemático: {json.dumps(legajo, ensure_ascii=False)[:500]}...")
                except Exception:
                    pass  # por si el legajo no es serializable

        # Resultados finales
        if resultados:
            # legajo_id puede ser str/int: normalizamos el sort por str para evitar TypeError
            resultados_ordenados = sorted(resultados, key=lambda x: (str(x[0]), x[1]))
            logger.info(
                f"✅ Proceso completado:\n"
                f"- Legajos procesados: {stats['legajos_procesados']}/{stats['total_legajos']}\n"
                f"- Variables calculadas: {stats['variables_calculadas']}\n"
                f"- Errores: {stats['legajos_con_error']}\n"
                f"- Tipos de errores: {dict(stats['errores_por_tipo'])}"
            )
            return resultados_ordenados, stats, resumen_horarios
        else:
            logger.warning("❌ No se generaron resultados válidos")
            return None, stats, resumen_horarios

    except json.JSONDecodeError as je:
        logger.error(f"El archivo no es un JSON válido: {str(je)}")
        return None, stats, resumen_horarios
    except FileNotFoundError:
        logger.error(f"Archivo no encontrado: {ruta_archivo}")
        return None, stats, resumen_horarios
    except Exception as e:
        logger.critical(f"Error inesperado: {str(e)}\n{traceback.format_exc()}")
        return None, stats, resumen_horarios
    
def guardar_resultados_csv(resultados: List[Tuple[int, int, Any]], nombre_archivo: str = 'variables_calculadas.xlsx') -> None:
    try:
        # Crear libro y hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Variables Calculadas"

        # Estilo encabezado
        encabezados = ['LEGAJO', 'CODIGO VARIABLE', 'VALOR']
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        for col_num, encabezado in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col_num, value=encabezado)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal='center')

        # Cuerpo del Excel
        fila_excel = 2
        for fila in resultados:
            if isinstance(fila, tuple) and len(fila) == 3:
                id_legajo, codigo_variable, valor = fila

                if isinstance(valor, (float, int)):
                    valor_str = f"{valor:.5f}".rstrip('0').rstrip('.').replace('.', ',')
                else:
                    valor_str = str(valor)

                ws.cell(row=fila_excel, column=1, value=id_legajo)
                ws.cell(row=fila_excel, column=2, value=codigo_variable)
                ws.cell(row=fila_excel, column=3, value=valor_str)
                fila_excel += 1
            else:
                logger.warning(f"Se encontró un resultado mal formado y fue omitido: {fila}")

        # Ajuste automático de ancho
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Guardar archivo
        nombre_archivo = os.path.join(os.getcwd(), nombre_archivo)
        wb.save(nombre_archivo)
        logger.info(f"✅ Archivo Excel guardado con formato visual en: {nombre_archivo}")

    except Exception as e:
        logger.error(f"❌ Error al guardar archivo Excel: {e}", exc_info=True)

def calcular_variables(legajo: Dict[str, Any]) -> List[Tuple[int, Any]]:
    """
    Calcula todas las variables para un legajo según las reglas establecidas.
    NUEVA VERSIÓN con logging estandarizado y completo.
    """
    variables = []
    id_legajo = legajo.get('id_legajo', 'ID_DESCONOCIDO_EN_CALCULO')
    
    try:
        logger.info(f"\n{'='*80}")
        logger.info(f"INICIANDO CÁLCULO - Legajo {id_legajo}")
        logger.info(f"{'='*80}\n")

        # ==========================================
        # VALIDACIÓN INICIAL
        # ==========================================
        if not validar_horario(legajo):
            logger.warning(f"Legajo {id_legajo}: Horario ambiguo/inválido")
            log_variable_calculada(id_legajo, 9000, "No se pudo interpretar correctamente el horario", 
                                  "Horario inválido")
            variables.append((9000, "No se pudo interpretar correctamente el horario"))
            log_resumen_variables(id_legajo, variables)
            return variables

        # ==========================================
        # VARIABLES BASE (FUNDACIONALES)
        # ==========================================
        
        # --- Variable 239: Horas Semanales ---
        log_variable_evaluando(id_legajo, 239)
        v239 = obtener_horas_semanales(legajo)
        variables.append((239, round(v239, 2)))
        log_variable_calculada(id_legajo, 239, round(v239, 2))

        # --- Variable 1242: Días Mensuales ---
        log_variable_evaluando(id_legajo, 1242)
        v1242 = calcular_dias_mensuales(legajo)
        variables.append((1242, v1242))
        log_variable_calculada(id_legajo, 1242, v1242)
        
        # --- Determinar si es guardia (no es variable, pero afecta cálculos) ---
        es_guardia_actual = es_guardia(legajo)
        logger.debug(f"Legajo {id_legajo}: es_guardia = {es_guardia_actual}")

        # ==========================================
        # VARIABLE 1: SUELDO BRUTO PACTADO
        # ==========================================
        log_variable_evaluando(id_legajo, 1)
        if cumple_condicion_sueldo_basico(legajo):
            sueldo = round(float(legajo.get('remuneracion', {}).get('sueldo_base', 0.0)), 2)
            variables.append((1, sueldo))
            log_variable_calculada(id_legajo, 1, sueldo)
        else:
            log_variable_no_calculada(id_legajo, 1, "No cumple condiciones de sueldo básico")

        # ==========================================
        # VARIABLE 2000: PERSONAL DE GUARDIA
        # ==========================================
        log_variable_evaluando(id_legajo, 2000)
        if es_guardia_actual:
            variables.append((2000, 1))
            log_variable_calculada(id_legajo, 2000, 1, "Es personal de guardia")
        else:
            log_variable_no_calculada(id_legajo, 2000, "No es personal de guardia")
        
        # ==========================================
        # VARIABLE 4: HORAS MENSUALES
        # ==========================================
        log_variable_evaluando(id_legajo, 4)
        v4 = calcular_horas_mensuales(legajo, v239)
        variables.append((4, round(v4, 2)))
        log_variable_calculada(id_legajo, 4, round(v4, 2))

        # ==========================================
        # VARIABLES 1157 y 1498: HORAS NOCTURNAS
        # ==========================================
        v1157 = obtener_horas_nocturnas(legajo, es_guardia_actual)
        full_nocturno = es_full_nocturno(legajo) if v1157 > 0 else False
        
        log_variable_evaluando(id_legajo, 1157)
        log_variable_evaluando(id_legajo, 1498)
        
        if v1157 == 0:
            log_variable_no_calculada(id_legajo, 1157, "Sin horas nocturnas")
            log_variable_no_calculada(id_legajo, 1498, "Sin horas nocturnas")
        elif full_nocturno:
            # CASO FULL NOCTURNO: Solo V1498
            log_variable_no_calculada(id_legajo, 1157, "Full nocturno - solo se liquida V1498")
            if aplicar_adicional_nocturno(legajo, v1157, es_guardia_actual):
                variables.append((1498, 1))
                log_variable_calculada(id_legajo, 1498, 1, "Full nocturno")
            else:
                log_variable_no_calculada(id_legajo, 1498, "No cumple condiciones de adicional nocturno")
        else:
            # CASO NORMAL
            variables.append((1157, round(v1157, 2)))
            log_variable_calculada(id_legajo, 1157, round(v1157, 2), f"{v1157} horas mensuales")
            
            if aplicar_adicional_nocturno(legajo, v1157, es_guardia_actual):
                variables.append((1498, 1))
                log_variable_calculada(id_legajo, 1498, 1)
            else:
                log_variable_no_calculada(id_legajo, 1498, "No cumple condiciones de adicional nocturno")

        # ==========================================
        # VARIABLE 992: EXTENSIÓN HORARIA
        # ==========================================
        log_variable_evaluando(id_legajo, 992)
        v992 = calcular_extension_horaria(legajo, v239)
        if v992 is not None:
            variables.append((992, round(v992, 2)))
            log_variable_calculada(id_legajo, 992, round(v992, 2))
        else:
            log_variable_no_calculada(id_legajo, 992, "No cumple condiciones")

        # ==========================================
        # VARIABLES 1145 y 1144: ADICIONAL PIVOT
        # ==========================================
        log_variable_evaluando(id_legajo, 1145)
        log_variable_evaluando(id_legajo, 1144)
        variables_pivot = calcular_adicional_pivot(legajo)

        v1145 = variables_pivot.get(1145)
        if v1145 is not None:
            variables.append((1145, v1145))
            log_variable_calculada(id_legajo, 1145, v1145, "Adicional pivot resonancia")
        else:
            log_variable_no_calculada(id_legajo, 1145, "No cumple condiciones del adicional pivot resonancia")

        v1144 = variables_pivot.get(1144)
        if v1144 is not None:
            variables.append((1144, v1144))
            log_variable_calculada(id_legajo, 1144, v1144, "Adicional pivot general")
        else:
            log_variable_no_calculada(id_legajo, 1144, "No cumple condiciones del adicional pivot general")

        # ==========================================
        # VARIABLE 1151: ADICIONAL RESONANCIA
        # ==========================================
        if VARIABLE_1151_HABILITADA:
            log_variable_evaluando(id_legajo, 1151)
            v1151 = calcular_adicional_resonancia(legajo, v239)
            if v1151 is not None:
                variables.append((1151, v1151))
                if isinstance(v1151, (int, float)):
                    log_variable_calculada(id_legajo, 1151, v1151)
                else:
                    log_variable_calculada(id_legajo, 1151, v1151, "Mensaje de validación")
            else:
                log_variable_no_calculada(id_legajo, 1151, "No cumple condiciones")
        else:
            log_variable_no_calculada(id_legajo, 1151, "Variable desactivada temporalmente")

        # ==========================================
        # VARIABLE 1131: DÍAS ESPECIALES
        # ==========================================
        log_variable_evaluando(id_legajo, 1131)
        v1131 = calcular_dias_especiales(legajo, v1242)
        if v1131 is not None:
            variables.append((1131, v1131))
            log_variable_calculada(id_legajo, 1131, v1131)
        else:
            log_variable_no_calculada(id_legajo, 1131, "No cumple condiciones")

        # ==========================================
        # VARIABLE 1137: LAVADO DE UNIFORME
        # ==========================================
        log_variable_evaluando(id_legajo, 1137)
        if aplicar_lavado_uniforme(legajo):
            variables.append((1137, 1))
            log_variable_calculada(id_legajo, 1137, 1)
        else:
            log_variable_no_calculada(id_legajo, 1137, "No cumple condiciones")

        # ==========================================
        # VARIABLE 1167: JORNADA REDUCIDA
        # ==========================================
        log_variable_evaluando(id_legajo, 1167)
        v1167 = calcular_jornada_reducida(legajo, es_guardia_actual)
        if v1167 is not None:
            variables.append((1167, v1167))
            log_variable_calculada(id_legajo, 1167, v1167, f"{v1167}%")
        else:
            log_variable_no_calculada(id_legajo, 1167, "No aplica jornada reducida")

        # ==========================================
        # VARIABLE 1416: JORNADA ART. 19
        # ==========================================
        log_variable_evaluando(id_legajo, 1416)
        v1416 = calcular_jornada_art19(legajo, v239)
        if v1416 is not None:
            variables.append((1416, v1416))
            log_variable_calculada(id_legajo, 1416, v1416)
        else:
            log_variable_no_calculada(id_legajo, 1416, "No cumple condiciones Art. 19")

        # ==========================================
        # VARIABLE 1599: PORCENTAJE ART. 19
        # ==========================================
        log_variable_evaluando(id_legajo, 1599)
        v1599 = calcular_porcentaje_art19(legajo, v239)
        if v1599 is not None:
            variables.append((1599, round(v1599, 4)))
            log_variable_calculada(id_legajo, 1599, round(v1599, 4), f"{v1599}%")
        else:
            log_variable_no_calculada(id_legajo, 1599, "No cumple condiciones Art. 19")

        # ==========================================
        # VARIABLE 1673: PROPORCIÓN LAVADO
        # ==========================================
        log_variable_evaluando(id_legajo, 1673)
        if aplicar_proporcion_lavado(legajo):
            variables.append((1673, 1))
            log_variable_calculada(id_legajo, 1673, 1)
        else:
            log_variable_no_calculada(id_legajo, 1673, "No cumple condiciones")

        # ==========================================
        # VARIABLE 2006: FECHA FIN CONTRATO
        # ==========================================
        log_variable_evaluando(id_legajo, 2006)
        fecha_fin = obtener_fecha_fin_contrato(legajo)
        if fecha_fin:
            variables.append((2006, fecha_fin))
            log_variable_calculada(id_legajo, 2006, fecha_fin)
        else:
            log_variable_no_calculada(id_legajo, 2006, "Sin fecha de fin de contrato")

        # ==========================================
        # VARIABLE 2281: NO LIQUIDA PLUS GUARDIA
        # ==========================================
        log_variable_evaluando(id_legajo, 2281)
        if aplicar_no_liquida_plus(legajo, es_guardia_actual):
            variables.append((2281, 1))
            log_variable_calculada(id_legajo, 2281, 1)
        else:
            log_variable_no_calculada(id_legajo, 2281, "No cumple condiciones")

        # ==========================================
        # VARIABLE 426: CAJERO/SEGURO
        # ==========================================
        log_variable_evaluando(id_legajo, 426)
        if es_cajero(legajo):
            variables.append((426, 1))
            log_variable_calculada(id_legajo, 426, 1)
        else:
            log_variable_no_calculada(id_legajo, 426, "No es cajero")

        # ==========================================
        # VARIABLES INFORMATIVAS (7000-13000)
        # ==========================================
        procesar_variables_informativas(legajo, variables)
        
        # ==========================================
        # VARIABLES MÉDICAS (1740, 1251, 1252)
        # ==========================================
        log_variable_evaluando(id_legajo, 1740)
        log_variable_evaluando(id_legajo, 1251)
        log_variable_evaluando(id_legajo, 1252)
        
        if es_medico_productividad(legajo):
            variables.extend([(1740, 1), (1251, 1), (1252, 1)])
            log_variable_calculada(id_legajo, 1740, 1, "Médico productividad")
            log_variable_calculada(id_legajo, 1251, 1, "Médico productividad")
            log_variable_calculada(id_legajo, 1252, 1, "Médico productividad")
        else:
            log_variable_no_calculada(id_legajo, 1740, "No es médico de productividad")
            log_variable_no_calculada(id_legajo, 1251, "No es médico de productividad")
            log_variable_no_calculada(id_legajo, 1252, "No es médico de productividad")

        # ==========================================
        # RESUMEN FINAL
        # ==========================================
        log_resumen_variables(id_legajo, variables)
        return variables

    except Exception as e:
        logger.error(f"ERROR CRÍTICO en legajo {id_legajo}: {str(e)}", exc_info=True)
        return []
    
# FUNCIONES DE VALIDACIÓN
# ==============================

def validar_estructura_legajo(legajo: Dict[str, Any]) -> bool:
    """Valida que el legajo tenga la estructura mínima requerida"""
    campos_requeridos = ['id_legajo', 'datos_personales', 'contratacion', 'horario', 'remuneracion']

    if not all(k in legajo for k in campos_requeridos):
        logger.warning(f"Legajo {legajo.get('id_legajo', 'DESCONOCIDO')} tiene estructura incompleta")
        return False

    subcampos_requeridos = [
        ('datos_personales', ['nombre', 'sector', 'puesto', 'sede']), 
        ('contratacion', ['tipo', 'categoria', 'fechas']),
        ('horario', ['bloques', 'resumen']),
        ('remuneracion', ['sueldo_base', 'moneda'])
    ]

    for campo, subcampos in subcampos_requeridos:
        if not all(k in legajo.get(campo, {}) for k in subcampos):
            logger.warning(f"Legajo {legajo['id_legajo']} no tiene todos los subcampos requeridos en {campo}")
            return False

    return True

def validar_horario(legajo: Dict[str, Any]) -> bool:
    """
    Valida si el horario es interpretable

    Args:
        legajo: Diccionario con datos del legajo

    Returns:
        True si el horario es válido, False si es ambiguo/inválido
    """
    if not legajo['horario']['bloques']:
        logger.warning(f"Legajo {legajo['id_legajo']}: Horario vacío")
        return False

    # Validación adicional de estructura de bloques horarios
    for bloque in legajo['horario']['bloques']:
        if not all(k in bloque for k in ['dias_semana', 'hora_inicio', 'hora_fin']):
            logger.warning(f"Legajo {legajo['id_legajo']}: Bloque horario incompleto")
            return False

    return True

def contiene_full_guardia(texto: str) -> bool:
    """
    Detecta 'full guardia' en cualquier formato con tolerancia a:
    - Mayúsculas/minúsculas
    - Espacios extras: 'full  guardia'
    - Guiones: 'full-guardia'
    - Typos menores: 'ful guardia', 'fullgardia'
    """
    if not texto or not isinstance(texto, str):
        return False
    
    texto_limpio = re.sub(r'[^\w\s-]', ' ', texto.lower())  # Elimina puntuación excepto guiones
    texto_limpio = re.sub(r'\s+', ' ', texto_limpio).strip()  # Normaliza espacios
    
    patron = re.compile(
        r'(?:full\s*[-]?\s*gu?a?rdia|gu?a?rdia\s*[-]?\s*full)',  # Admite orden invertido
        re.IGNORECASE
    )
    return bool(patron.search(texto_limpio))

def es_guardia(legajo: Dict[str, Any]) -> bool:
    """
    Determina si un legajo es GUARDIA según 2 condiciones acumulativas:
    1) Sede válida (según lista normalizada)
    2) Contiene 'full guardia' en adicionables
    """
    try:
        id_legajo = legajo.get('id_legajo', 'N/A')
        sede_raw = legajo.get('datos_personales', {}).get('sede', '')
        sede_normalizada = normalizar_texto(sede_raw)

        sede_valida = sede_normalizada in sedes_permitidas
        logger.debug(f"[es_guardia] Legajo {id_legajo}: Sede normalizada = '{sede_normalizada}', válida = {sede_valida}")
        if not sede_valida:
            logger.debug(f"[es_guardia] Legajo {id_legajo}: Sede '{sede_raw}' NO válida.")
            return False

        # --- 2. Validación de Adicionables ---
        adicionables = str(legajo.get('remuneracion', {}).get('adicionables') or '')
        adicionables_normalizados = normalizar_texto(adicionables)

        if 'full guardia' not in adicionables_normalizados:
            logger.debug(f"[es_guardia] Legajo {id_legajo}: Adicionables NO contienen 'full guardia'.")
            return False

        # --- Pasa TODAS las condiciones ---
        logger.info(f"[es_guardia] Legajo {id_legajo}: ✅ Validado como GUARDIA (sede='{sede_raw}')")
        return True

    except Exception as e:
        logger.error(f"[es_guardia] Legajo {legajo.get('id_legajo', 'N/A')}: ❌ Error inesperado - {str(e)}")
        logger.error(traceback.format_exc())
        return False

    # 1. Helper function adaptada para el formato de tus constantes
def es_puesto_especial(puesto_normalizado: str) -> bool:
    """Versión mejorada para evitar falsos positivos"""
    # Limpieza adicional
    puesto_limpio = re.sub(r'\s+\bde\b\s+', ' ', puesto_normalizado).strip().lower()
    puesto_limpio = re.sub(r'[^a-z0-9 ]', '', puesto_limpio)  # Elimina caracteres especiales
    
    # Comparación más estricta
    for puesto_especial in PUESTOS_ESPECIALES.values():
        especial_limpio = re.sub(r'\s+\bde\b\s+', ' ', puesto_especial).strip().lower()
        especial_limpio = re.sub(r'[^a-z0-9 ]', '', especial_limpio)
        
        # Coincidencia exacta o comienzo del string
        if (puesto_limpio == especial_limpio or 
            puesto_limpio.startswith(especial_limpio + " ") or 
            especial_limpio.startswith(puesto_limpio + " ")):
            return True
            
    return False

def _parse_fecha_flexible(valor: Any) -> Optional[datetime]:
    """
    Intenta parsear una fecha en múltiples formatos comunes.
    Soporta:
      - Separadores: '/', '-', '.'
      - Años de 2 o 4 dígitos (25 -> 2025 por %y)
      - Ordenes habituales: dd/mm/aa(aa), dd-mm-aa(aa), aa(aa)-mm-dd, aa(aa)/mm/dd, dd.mm.aa(aa)
    Retorna un datetime o None si no pudo parsear.
    """
    logger = logging.getLogger(__name__)

    if valor is None:
        return None

    # Normalizar a str y limpiar espacios
    s = str(valor).strip()
    if not s:
        return None

    # Normalizamos unicode (por si viene con caracteres raros)
    s = unicodedata.normalize("NFKC", s)

    # Cambiamos cualquier separador no numérico por '/'
    s_norm = re.sub(r"[^0-9]", "/", s)
    s_norm = re.sub(r"/+", "/", s_norm).strip("/")

    # Lista de formatos a probar (dos y cuatro dígitos de año)
    formatos = [
        "%d/%m/%Y", "%d/%m/%y",
        "%d-%m-%Y", "%d-%m-%y",  # por si el usuario no normalizó separadores
        "%Y/%m/%d", "%y/%m/%d",
        "%Y-%m-%d", "%y-%m-%d",
        "%d.%m.%Y", "%d.%m.%y",
    ]

    # Primero probamos con la cadena original y sus variantes normalizadas
    candidatos = {s, s_norm, s_norm.replace("/", "-"), s_norm.replace("/", ".")}

    for cand in list(candidatos):
        for fmt in formatos:
            try:
                # Si el formato usa '-' o '.' lo probamos también
                cand_fmt = cand
                if "." in fmt:
                    cand_fmt = cand.replace("/", ".")
                elif "-" in fmt:
                    cand_fmt = cand.replace("/", "-")
                else:
                    cand_fmt = cand.replace("-", "/").replace(".", "/")

                dt = datetime.strptime(cand_fmt, fmt)
                # %y mapea 00-68 a 2000-2068 -> "25" => 2025 (justo lo que queremos)
                return dt
            except ValueError:
                continue

    # Heurística extra: si tenemos exactamente 3 grupos numéricos, intentamos reordenar
    partes = re.split(r"[^\d]", s)
    partes = [p for p in partes if p.isdigit()]
    if len(partes) == 3:
        d, m, a = partes[0], partes[1], partes[2]
        # Intento dd/mm/aa(aa)
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(f"{d}/{m}/{a}", fmt)
            except ValueError:
                pass
        # Intento aa(aa)/mm/dd
        for fmt in ("%Y/%m/%d", "%y/%m/%d"):
            try:
                return datetime.strptime(f"{a}/{m}/{d}", fmt)
            except ValueError:
                pass

    logger.debug(f"_parse_fecha_flexible: no se pudo interpretar la fecha '{valor}'")
    return None

# ==============================
# FUNCIONES DE CÁLCULO
# ==============================

def obtener_horas_semanales(legajo: Dict[str, Any]) -> float:
    try:
        # Uso robusto de .get()
        horas_raw = legajo.get('horario', {}).get('resumen', {}).get('total_horas_semanales')

        if horas_raw is None:
            logger.warning(f"Legajo {legajo.get('id_legajo', 'N/A')}: 'total_horas_semanales' es None. Devolviendo 0.0.")
            return 0.0

        horas = float(horas_raw)
        if horas < 0 or horas > 168:
            logger.warning(f"Legajo {legajo.get('id_legajo', 'N/A')}: Horas semanales fuera de rango ({horas})")
            return 0.0
        return horas
    except (TypeError, ValueError) as e: # KeyError ya no es probable con .get()
        logger.error(f"Legajo {legajo.get('id_legajo', 'N/A')}: Error al convertir horas semanales a float - {str(e)}")
        return 0.0
    except Exception as e: # Para cualquier otro error inesperado
        logger.error(f"Legajo {legajo.get('id_legajo', 'N/A')}: Error inesperado al obtener horas semanales - {str(e)}")
        logger.error(traceback.format_exc())
        return 0.0

def calcular_dias_mensuales(legajo: Dict[str, Any]) -> int:
    """
    Calcula días mensuales ajustando correctamente días con periodicidad quincenal o parcial.
    Versión corregida: procesa correctamente todos los bloques por día.
    """
    id_legajo = legajo.get("id_legajo", "DESCONOCIDO")

    try:
        bloques_por_dia = legajo.get("horario", {}).get("resumen", {}).get("bloques_por_dia", {})

        if not isinstance(bloques_por_dia, dict) or not bloques_por_dia:
            logger.warning(f"Legajo {id_legajo}: 'bloques_por_dia' ausente o vacío.")
            return 0

        dias_semanales = 0.0

        for dia_str, bloques in bloques_por_dia.items():
            if not isinstance(bloques, list) or not bloques:
                continue

            dia_procesado = False

            for bloque in bloques:
                if not isinstance(bloque, dict):
                    continue
                    
                periodicidad = str(bloque.get("periodicidad", "")).lower()
                
                if periodicidad == "semanal" and not dia_procesado:
                    dias_semanales += 1.0
                    dia_procesado = True
                    logger.debug(f"Legajo {id_legajo}: Día {dia_str} → semanal (1.0)")
                    
                elif periodicidad == "quincenal" and not dia_procesado:
                    dias_semanales += 0.5
                    dia_procesado = True
                    logger.debug(f"Legajo {id_legajo}: Día {dia_str} → quincenal (0.5)")

                # ===== INICIO DE LA CORRECCIÓN =====
                elif periodicidad == "mensual" and not dia_procesado:
                    dias_semanales += 0.25  # 1 día al mes = 1/4 de día a la semana
                    dia_procesado = True
                    logger.debug(f"Legajo {id_legajo}: Día {dia_str} → mensual (0.25)")
                # ===== FIN DE LA CORRECCIÓN =====
                    
                elif periodicidad == "proporcional" and not dia_procesado:
                    # CALCULAR FACTOR PROPORCIONAL
                    horas_semanales = bloque.get("horas_semanales", 0)
                    duracion_total = bloque.get("duracion_total", 1)
                    
                    if duracion_total > 0 and horas_semanales > 0:
                        factor = horas_semanales / duracion_total
                    else:
                        factor = 0.75  # Default
                    
                    dias_semanales += factor
                    dia_procesado = True
                    logger.debug(f"Legajo {id_legajo}: Día {dia_str} → proporcional (factor {factor})")

            # Si no se procesó el día (sin periodicidad reconocida), contar como semanal
            if not dia_procesado:
                dias_semanales += 1.0
                logger.debug(f"Legajo {id_legajo}: Día {dia_str} → sin periodicidad (1.0)")

        dias_mensuales = dias_semanales * 4.33
        # Usamos un redondeo estándar (ej: 22.7 -> 23)
        dias_mensuales_redondeados = int(dias_mensuales + 0.5)

        logger.info(
            f"Legajo {id_legajo}: Días semanales efectivos = {dias_semanales:.2f}, "
            f"mensuales estimados = {dias_mensuales:.2f}, redondeado = {dias_mensuales_redondeados}"
        )

        return dias_mensuales_redondeados

    except Exception as e:
        logger.error(f"Legajo {id_legajo}: Error al calcular días mensuales. Detalle: {str(e)}")
        # import traceback; logger.error(traceback.format_exc()) # Descomentar para debug más profundo
        return 0
    
def cumple_condicion_sueldo_basico(legajo: Dict[str, Any]) -> bool:
    """
    Determina si aplica el sueldo básico (Variable 1) de forma robusta.
    Condiciones:
    1. Categoría debe ser 'fc_pfc'
    2. Debe tener sueldo_base válido (no None)
    3. sueldo_base debe ser convertible a número
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # 1. Validar categoría
        categoria = legajo.get('contratacion', {}).get('categoria')
        
        if categoria != 'fc_pfc':
            logger.debug(f"[V1] Legajo {id_legajo}: ✗ Categoría '{categoria}' != 'fc_pfc'")
            return False

        # 2. Validar sueldo_base existe
        sueldo = legajo.get('remuneracion', {}).get('sueldo_base')
        
        if sueldo is None:
            logger.debug(f"[V1] Legajo {id_legajo}: ✗ Sueldo base es None")
            return False

        # 3. Validar que sea numérico
        float(sueldo)  # Solo para validar
        return True

    except (KeyError, ValueError, TypeError) as e:
        logger.debug(f"[V1] Legajo {id_legajo}: ✗ Error: {str(e)}")
        return False

def es_full_nocturno(legajo: Dict[str, Any]) -> bool:
    """
    Determina si un legajo es "full nocturno" según 3 condiciones acumulativas:
    a) Más del 80% de los días tienen horario nocturno
    b) Más del 50% de las horas de cada jornada diaria son nocturnas
    c) La jornada debe comenzar a las 18:00 o después
    
    Si cumple TODAS las condiciones, se considera full nocturno y solo se liquida
    la variable 1498, NO la 1157.
    
    Args:
        legajo: Diccionario con datos del legajo
        
    Returns:
        bool: True si es full nocturno, False en caso contrario
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        resumen = legajo.get('horario', {}).get('resumen', {})
        bloques_por_dia = resumen.get('bloques_por_dia', {})
        
        if not bloques_por_dia:
            logger.debug(f"[full_nocturno] Legajo {id_legajo}: Sin bloques por día")
            return False
        
        total_dias = len(bloques_por_dia)
        if total_dias == 0:
            return False
        
        # Contadores
        dias_con_nocturnidad = 0
        dias_con_mayoria_nocturna = 0
        dias_con_inicio_18_o_despues = 0
        
        for dia_str, bloques_del_dia in bloques_por_dia.items():
            if not isinstance(bloques_del_dia, list) or len(bloques_del_dia) == 0:
                continue
            
            # Variables para este día
            tiene_horas_nocturnas = False
            total_horas_dia = 0.0
            total_horas_nocturnas_dia = 0.0
            hora_inicio_mas_temprana = None
            
            for bloque in bloques_del_dia:
                if not isinstance(bloque, dict):
                    continue
                
                duracion = bloque.get('duracion_total', 0)
                horas_noct = bloque.get('horas_nocturnas', 0)
                inicio = bloque.get('inicio', '')
                
                total_horas_dia += duracion
                total_horas_nocturnas_dia += horas_noct
                
                if horas_noct > 0:
                    tiene_horas_nocturnas = True
                
                # Obtener la hora de inicio más temprana del día
                if inicio:
                    try:
                        # Convertir "18:00" a minutos desde medianoche para comparar
                        partes = inicio.split(':')
                        if len(partes) == 2:
                            minutos_inicio = int(partes[0]) * 60 + int(partes[1])
                            if hora_inicio_mas_temprana is None or minutos_inicio < hora_inicio_mas_temprana:
                                hora_inicio_mas_temprana = minutos_inicio
                    except (ValueError, IndexError):
                        pass
            
            # Condición a) ¿Este día tiene horario nocturno?
            if tiene_horas_nocturnas:
                dias_con_nocturnidad += 1
            
            # Condición b) ¿Más del 50% de las horas de este día son nocturnas?
            if total_horas_dia > 0 and (total_horas_nocturnas_dia / total_horas_dia) > 0.5:
                dias_con_mayoria_nocturna += 1
            
            # Condición c) ¿La jornada comienza a las 18:00 o después?
            # 18:00 = 18 * 60 = 1080 minutos
            if hora_inicio_mas_temprana is not None and hora_inicio_mas_temprana >= 1080:
                dias_con_inicio_18_o_despues += 1
        
        # Calcular porcentajes
        porcentaje_dias_nocturnos = (dias_con_nocturnidad / total_dias) * 100 if total_dias > 0 else 0
        
        logger.debug(
            f"[full_nocturno] Legajo {id_legajo}: "
            f"Total días={total_dias}, "
            f"Días con nocturnidad={dias_con_nocturnidad} ({porcentaje_dias_nocturnos:.1f}%), "
            f"Días con mayoría nocturna={dias_con_mayoria_nocturna}, "
            f"Días inicio >=18:00={dias_con_inicio_18_o_despues}"
        )
        
        # Evaluar las 3 condiciones
        condicion_a = porcentaje_dias_nocturnos > 80
        condicion_b = dias_con_mayoria_nocturna == total_dias  # TODOS los días deben tener mayoría nocturna
        condicion_c = dias_con_inicio_18_o_despues == total_dias  # TODOS los días deben iniciar >= 18:00
        
        es_full = condicion_a and condicion_b and condicion_c
        
        if es_full:
            logger.info(
                f"[full_nocturno] Legajo {id_legajo}: ✅ ES FULL NOCTURNO "
                f"(a={condicion_a}, b={condicion_b}, c={condicion_c})"
            )
        else:
            logger.debug(
                f"[full_nocturno] Legajo {id_legajo}: NO es full nocturno "
                f"(a={condicion_a}, b={condicion_b}, c={condicion_c})"
            )
        
        return es_full
        
    except Exception as e:
        logger.error(f"[full_nocturno] Legajo {id_legajo}: Error - {str(e)}")
        logger.error(traceback.format_exc())
        return False

def obtener_horas_nocturnas(legajo: Dict[str, Any], es_guardia: bool) -> float:
    """
    Calcula horas nocturnas MENSUALES válidas para un legajo, considerando:
    - Guardias: siempre retorna 0.0
    - No guardias: horas semanales × 4.33 (conversión a mensual)
    
    Args:
        legajo: Diccionario con datos del legajo
        es_guardia: Resultado de la función es_guardia()
        
    Returns:
        float: Horas nocturnas MENSUALES (horas_semanales × 4.33)
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    # 1. Guardias no acumulan horas nocturnas
    if es_guardia:
        logger.debug(f"[V1157] Legajo {id_legajo}: ✗ Es guardia → horas nocturnas=0")
        return 0.0
    
    try:
        # 2. Obtener y validar horas semanales de forma robusta
        horas_semanales_raw = legajo.get('horario', {}).get('resumen', {}).get('total_horas_nocturnas', 0)
        
        logger.debug(f"[V1157] Legajo {id_legajo}: ✓ Horas nocturnas semanales raw={horas_semanales_raw}")
        
        horas_semanales = float(horas_semanales_raw)
        
        # 3. Aplicar límites razonables (0 <= horas <= 168)
        horas_semanales_validas = max(0.0, min(horas_semanales, 168.0))
        
        if abs(horas_semanales_validas - horas_semanales) > 0.01:  # Tolerancia para floats
            logger.warning(f"[V1157] Legajo {id_legajo}: ⚠ Ajustadas horas {horas_semanales} → {horas_semanales_validas}")
        
        # 4. MULTIPLICAR POR 4.33 para obtener horas mensuales
        horas_mensuales = round(horas_semanales_validas * 4.33, 2)
        
        logger.debug(f"[V1157] Legajo {id_legajo}: ✓ Semanales={horas_semanales_validas} → Mensuales (×4.33)={horas_mensuales}")
        
        if horas_mensuales > 0:
            logger.info(f"[V1157] Legajo {id_legajo}: ✓ RESULTADO = {horas_mensuales} horas")
        else:
            logger.debug(f"[V1157] Legajo {id_legajo}: ✗ Sin horas nocturnas")
        
        return horas_mensuales
        
    except (TypeError, ValueError) as e:
        logger.error(f"[V1157] Legajo {id_legajo}: ERROR - Valor inválido - {str(e)}")
        return 0.0
    except Exception as e:
        logger.error(f"[V1157] Legajo {id_legajo}: ERROR CRÍTICO - {str(e)}")
        logger.error(traceback.format_exc())
        return 0.0
    
def aplicar_lavado_uniforme(legajo: Dict[str, Any]) -> bool:
    """
    Determina si aplica lavado de uniforme (Variable 1137).
    Condiciones:
    - Puesto: OPERARIO DE LOGISTICA
    - Subsector: INTERIOR
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # Acceder a datos_personales
        datos_personales = legajo.get('datos_personales')
        if not isinstance(datos_personales, dict):
            logger.debug(f"[V1137] Legajo {id_legajo}: ✗ datos_personales inválido")
            return False

        # Normalizar puesto
        puesto_raw = datos_personales.get('puesto')
        puesto_normalizado = normalizar_texto(puesto_raw)

        # Acceder a sector
        sector_data = datos_personales.get('sector')
        if not isinstance(sector_data, dict):
            logger.debug(f"[V1137] Legajo {id_legajo}: ✗ Sector inválido")
            return False

        # Normalizar subsector
        subsector_raw = sector_data.get('subsector')
        subsector_normalizado = normalizar_texto(subsector_raw)

        # Validar condiciones
        puesto_ok = puesto_normalizado == normalizar_texto("OPERARIO DE LOGISTICA")
        subsector_ok = subsector_normalizado == normalizar_texto("INTERIOR")
        
        resultado = puesto_ok and subsector_ok
        if not resultado:
            logger.debug(f"[V1137] Legajo {id_legajo}: ✗ Puesto='{puesto_raw}', Subsector='{subsector_raw}'")
        
        return resultado

    except KeyError as ke:
        logger.error(f"Legajo {legajo.get('id_legajo', 'UNKNOWN')}: Falta clave esencial para validar lavado de uniforme - {str(ke)}")
        return False
    except Exception as e:
        logger.error(f"Legajo {legajo.get('id_legajo', 'UNKNOWN')}: Error general validando lavado de uniforme - {str(e)}")
        logger.error(traceback.format_exc())
        return False

def aplicar_adicional_nocturno(legajo: Dict[str, Any], horas_nocturnas: float, es_guardia: bool) -> bool:
    """
    Determina si aplica adicional nocturno según:
    1) NO sea guardia
    2) Tenga horas nocturnas > 0
    3) Pertenezca a categoría DC (Dentro de Convenio)
    Args:
        legajo: Diccionario con datos del legajo
        horas_nocturnas: Horas calculadas por obtener_horas_nocturnas()
        es_guardia: Resultado de es_guardia()
    Returns:
        bool: True si cumple TODAS las condiciones
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    logger.debug(f"[V1498] Legajo {id_legajo}: INICIO EVALUACIÓN")
    logger.debug(f"[V1498] Legajo {id_legajo}:   - es_guardia={es_guardia}")
    logger.debug(f"[V1498] Legajo {id_legajo}:   - horas_nocturnas={horas_nocturnas}")

    # 1. Excepciones rápidas (guardias o sin horas nocturnas)
    if es_guardia:
        logger.debug(f"[V1498] Legajo {id_legajo}: ✗ Excluido (es guardia)")
        return False
    if horas_nocturnas <= 0:
        logger.debug(f"[V1498] Legajo {id_legajo}: ✗ Excluido (sin horas nocturnas)")
        return False

    try:
        # 2. Validar categoría
        categoria = legajo.get('contratacion', {}).get('categoria', '')
        
        logger.debug(f"[V1498] Legajo {id_legajo}:   - Categoría='{categoria}'")
        
        if not categoria:
            logger.warning(f"[V1498] Legajo {id_legajo}: ⚠ Categoría vacía")
            return False
            
        # 3. Verificar convenio (DC = Dentro de Convenio)
        es_dc = str(categoria).lower().startswith('dc_')
        
        logger.debug(f"[V1498] Legajo {id_legajo}:   - ¿Empieza con 'dc_'?: {es_dc}")
        
        if es_dc:
            logger.info(f"[V1498] Legajo {id_legajo}: ✓ APLICA (DC, {horas_nocturnas}h)")
        else:
            logger.debug(f"[V1498] Legajo {id_legajo}: ✗ NO APLICA (Categoría '{categoria}' no es DC)")
        
        return es_dc
        
    except Exception as e:
        logger.error(f"[V1498] Legajo {id_legajo}: ERROR CRÍTICO - {str(e)}")
        logger.error(traceback.format_exc())
        return False

def obtener_fecha_fin_contrato(legajo: Dict[str, Any]) -> Optional[str]:
    """
    Obtiene fecha de fin de contrato para contratos a plazo fijo/determinado (Variable 2006).
    
    Condiciones acumulativas:
    - Tipo de contrato contiene 'plazo_fijo' o 'determinado'
    - Fecha fin es parseable
    
    Args:
        legajo: Diccionario con datos del legajo
        
    Returns:
        str | None: Fecha en formato dd/mm/YYYY o None si no aplica
    """
    logger = logging.getLogger(__name__)
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # 1. Obtener datos de contratación
        contratacion = legajo.get("contratacion", {}) or {}
        tipo_contrato_raw = contratacion.get("tipo", "") or ""
        tipo_contrato = str(tipo_contrato_raw).lower()
        logger.debug(f"[V2006] Legajo {id_legajo}: Tipo contrato = '{tipo_contrato_raw}'")
        
        # 2. Verificar si es plazo fijo/determinado
        es_plazo_fijo = any(t in tipo_contrato for t in ("plazo_fijo", "determinado"))
        logger.debug(f"[V2006] Legajo {id_legajo}: ¿Es plazo fijo/determinado? {es_plazo_fijo}")
        
        if not es_plazo_fijo:
            logger.debug(f"[V2006] Legajo {id_legajo}: ✗ NO APLICA - Tipo '{tipo_contrato_raw}' no es plazo fijo")
            return None
        
        # 3. Obtener fecha fin
        fechas = contratacion.get("fechas", {}) or {}
        fecha_fin_raw = fechas.get("fin")
        logger.debug(f"[V2006] Legajo {id_legajo}: Fecha fin raw = '{fecha_fin_raw}'")
        
        if not fecha_fin_raw:
            logger.debug(f"[V2006] Legajo {id_legajo}: ✗ NO APLICA - Fecha fin vacía/None")
            return None
        
        # 4. Parsear fecha
        fecha_obj = _parse_fecha_flexible(fecha_fin_raw)
        if not fecha_obj:
            logger.warning(f"[V2006] Legajo {id_legajo}: ✗ NO APLICA - No se pudo parsear fecha '{fecha_fin_raw}'")
            return None
        
        fecha_formateada = fecha_obj.strftime("%d/%m/%Y")
        logger.debug(f"[V2006] Legajo {id_legajo}: ✓ APLICA - Fecha fin = {fecha_formateada}")
        
        return fecha_formateada

    except Exception as e:
        logger.error(f"[V2006] Legajo {id_legajo}: Error obteniendo fecha fin contrato - {e}", exc_info=True)
        return None

def aplicar_no_liquida_plus(legajo: Dict[str, Any], es_guardia: bool) -> bool:
    """
    Determina si un legajo no debe liquidar plus (Variable 2281).
    Condiciones para NO liquidar:
    - No es guardia O
    - Legajo <= 15000 O
    - Pertenece a sedes excluidas (normalizado)
    
    Args:
        legajo: Diccionario con datos del legajo
        es_guardia: Booleano que indica si es guardia
        
    Returns:
        bool: True si NO debe liquidar plus, False si sí debe
    """
    id_legajo = legajo.get('id_legajo', 0)
    
    # 1. Validación: No es guardia
    if not es_guardia:
        logger.debug(f"[V2281] Legajo {id_legajo}: NO APLICA - No es guardia")
        return False
    
    # 2. Validación: Legajo <= 15000
    if id_legajo <= 15000:
        logger.debug(f"[V2281] Legajo {id_legajo}: NO APLICA - ID <= 15000")
        return False
    
    # 3. Obtener sede normalizada
    try:
        sede_actual = legajo.get('datos_personales', {}).get('sede')
        if not sede_actual:
            logger.debug(f"[V2281] Legajo {id_legajo}: NO APLICA - Sede no definida")
            return False
        
        sede_normalizada = normalizar_texto(sede_actual)
        logger.debug(f"[V2281] Legajo {id_legajo}: Sede = '{sede_actual}' (normalizado: '{sede_normalizada}')")
        
        # 4. Verificar si está en sedes excluidas
        en_lista_excluida = sede_normalizada in SEDES_NO_LIQUIDA_PLUS
        logger.debug(f"[V2281] Legajo {id_legajo}: ¿Sede en lista excluida? {en_lista_excluida}")
        
        if en_lista_excluida:
            logger.debug(f"[V2281] Legajo {id_legajo}: ✓ APLICA - Sede '{sede_actual}' NO liquida plus")
        else:
            logger.debug(f"[V2281] Legajo {id_legajo}: NO APLICA - Sede '{sede_actual}' SÍ liquida plus")
        
        return en_lista_excluida
        
    except Exception as e:
        logger.error(f"[V2281] Error en legajo {id_legajo}: {str(e)}")
        logger.error(traceback.format_exc())
        return False  # Por defecto, no aplicar restricción si hay error

def es_cajero(legajo: Dict[str, Any]) -> bool:
    """
    Determina si el legajo cumple criterios de cajero (Variable 426).
    Condiciones acumulativas:
    - Puesto contiene "CAJERO" o "CAJERO/A" (case-insensitive)
    - Categoría contiene "adm" o "administrativo" (normalizado)
    
    Args:
        legajo: Diccionario con datos del legajo
        
    Returns:
        bool: True si cumple criterios de cajero, False en caso contrario
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # 1. Obtener y normalizar puesto
        puesto_raw = legajo.get('datos_personales', {}).get('puesto')
        if not puesto_raw:
            logger.debug(f"[V426] Legajo {id_legajo}: ✗ NO APLICA - Puesto vacío/None")
            return False
        
        puesto = normalizar_texto(puesto_raw)
        logger.debug(f"[V426] Legajo {id_legajo}: Puesto = '{puesto_raw}' (normalizado: '{puesto}')")
        
        # 2. Verificar si puesto contiene "CAJERO" o "CAJERO/A"
        puesto_upper = puesto.upper()
        es_puesto_cajero = "CAJERO" in puesto_upper or "CAJERO/A" in puesto_upper
        logger.debug(f"[V426] Legajo {id_legajo}: ¿Puesto contiene CAJERO? {es_puesto_cajero}")
        
        if not es_puesto_cajero:
            logger.debug(f"[V426] Legajo {id_legajo}: ✗ NO APLICA - Puesto no es CAJERO")
            return False
        
        # 3. Obtener y normalizar categoría
        categoria_raw = legajo.get('contratacion', {}).get('categoria')
        if not categoria_raw:
            logger.debug(f"[V426] Legajo {id_legajo}: ✗ NO APLICA - Categoría vacía/None")
            return False
        
        categoria = normalizar_texto(categoria_raw)
        logger.debug(f"[V426] Legajo {id_legajo}: Categoría = '{categoria_raw}' (normalizado: '{categoria}')")
        
        # 4. Verificar si categoría contiene "adm" o "administrativo"
        es_categoria_adm = any(adm in categoria for adm in ['adm', 'administrativo'])
        logger.debug(f"[V426] Legajo {id_legajo}: ¿Categoría contiene 'adm'/'administrativo'? {es_categoria_adm}")
        
        if es_categoria_adm:
            logger.debug(f"[V426] Legajo {id_legajo}: ✓ APLICA - Cajero administrativo")
        else:
            logger.debug(f"[V426] Legajo {id_legajo}: ✗ NO APLICA - Categoría no es administrativa")
        
        return es_categoria_adm
        
    except KeyError as ke:
        logger.error(f"[V426] Legajo {id_legajo}: Falta clave en datos - {str(ke)}")
        return False
    except Exception as e:
        logger.error(f"[V426] Legajo {id_legajo}: Error validando cajero - {str(e)}")
        logger.error(traceback.format_exc())
        return False

def procesar_variables_informativas(legajo: Dict[str, Any], variables: List[Tuple[int, Any]]) -> None:
    """
    Procesa todas las variables informativas (7000-13000) con logging estandarizado.
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # Obtener adicionables normalizado
        adicionables_raw = legajo.get('remuneracion', {}).get('adicionables', '')
        adicionables_normalizado = normalizar_texto(adicionables_raw) if adicionables_raw else ""
        
        # Aplicar reemplazos específicos para 'intangibilidad'
        adicionables_para_intang = (adicionables_normalizado
                                    .replace("intang", "intangibilidad")
                                    .replace("intang.", "intangibilidad")
                                    .replace("intan", "intangibilidad")
                                    .replace("intangib", "intangibilidad"))

        sueldo_base = legajo.get('remuneracion', {}).get('sueldo_base')
        categoria = (legajo.get('contratacion', {}).get('categoria') or '').strip().lower()
        remuneracion = legajo.get('remuneracion', {})

        # ==========================================
        # VARIABLE 7000: CESIÓN
        # ==========================================
        log_variable_evaluando(id_legajo, 7000)
        if any(term in adicionables_normalizado for term in TERMINOS_CESION):
            variables.append((7000, "Es cesión, revisar."))
            log_variable_calculada(id_legajo, 7000, "Es cesión, revisar.")
        else:
            log_variable_no_calculada(id_legajo, 7000, "No contiene términos de cesión")

        # ==========================================
        # VARIABLE 8000: INTANGIBILIDAD
        # ==========================================
        log_variable_evaluando(id_legajo, 8000)
        if "intangibilidad" in adicionables_para_intang:
            variables.append((8000, "Revisar Importe o % para Intangibilidad Salarial"))
            log_variable_calculada(id_legajo, 8000, "Revisar Importe o % para Intangibilidad Salarial")
        else:
            log_variable_no_calculada(id_legajo, 8000, "No contiene intangibilidad")

        # ==========================================
        # VARIABLE 9000: ADICIONAL VOLUNTARIO
        # ==========================================
        log_variable_evaluando(id_legajo, 9000)
        terminos_adic_voluntario = ["adic voluntario", "adicional voluntario", "voluntario empresa"]
        if any(term in adicionables_normalizado for term in terminos_adic_voluntario):
            variables.append((9000, "Revisar Adic Voluntario Empresa"))
            log_variable_calculada(id_legajo, 9000, "Revisar Adic Voluntario Empresa")
        else:
            log_variable_no_calculada(id_legajo, 9000, "No contiene adicional voluntario")

        # ==========================================
        # VARIABLE 10000: LICENCIADO BIOIMÁGENES
        # ==========================================
        log_variable_evaluando(id_legajo, 10000)
        if es_licenciado_bioimagenes(legajo):
            variables.append((10000, "Cargar Título en CP, es Licenciado"))
            log_variable_calculada(id_legajo, 10000, "Cargar Título en CP, es Licenciado")
        else:
            log_variable_no_calculada(id_legajo, 10000, "No es licenciado en bioimágenes")

        # ==========================================
        # VARIABLE 11000: PPR
        # ==========================================
        log_variable_evaluando(id_legajo, 11000)
        ppr_en_adicionables = "ppr" in adicionables_normalizado
        sueldo_base_tiene_valor = sueldo_base is not None
        
        if ppr_en_adicionables and sueldo_base_tiene_valor:
            variables.append((11000, "Tiene PPR. Revisar archivo"))
            log_variable_calculada(id_legajo, 11000, "Tiene PPR. Revisar archivo")
        else:
            razon = "No tiene PPR en adicionables" if not ppr_en_adicionables else "Sin sueldo base"
            log_variable_no_calculada(id_legajo, 11000, razon)

        # ==========================================
        # VARIABLE 12000: FALTA SUELDO BRUTO PFC
        # ==========================================
        log_variable_evaluando(id_legajo, 12000)
        if categoria == "fc_pfc":
            sueldo_base_falta = (not isinstance(remuneracion, dict) or
                                 ('sueldo_base' not in remuneracion) or
                                 remuneracion.get('sueldo_base') in (None, ""))
            tiene_full_guardia = "full guardia" in adicionables_normalizado

            if sueldo_base_falta and not tiene_full_guardia:
                variables.append((12000, "Falta sueldo bruto pactado. Revisar Var 1"))
                log_variable_calculada(id_legajo, 12000, "Falta sueldo bruto pactado. Revisar Var 1")
            else:
                razon = "Tiene sueldo base" if not sueldo_base_falta else "Tiene full guardia"
                log_variable_no_calculada(id_legajo, 12000, razon)
        else:
            log_variable_no_calculada(id_legajo, 12000, "No es categoría FC_PFC")

        # ==========================================
        # VARIABLE 13000: GUARDIAS DE CAPACITACIÓN
        # ==========================================
        log_variable_evaluando(id_legajo, 13000)
        tiene_full_guardia = "full guardia" in adicionables_normalizado
        tiene_capacitacion = any(term in adicionables_normalizado for term in ["capacitacion", "capa"])
        
        if tiene_full_guardia and tiene_capacitacion:
            variables.append((13000, "Revisar Pago de Guardias de Capacitación"))
            log_variable_calculada(id_legajo, 13000, "Revisar Pago de Guardias de Capacitación")
        else:
            razon = "No tiene full guardia" if not tiene_full_guardia else "No tiene capacitación"
            log_variable_no_calculada(id_legajo, 13000, razon)

    except Exception as e:
        logger.error(f"{COLOR_BOLD}{COLOR_RED}Legajo {id_legajo}: Error procesando variables informativas - {str(e)}{COLOR_RESET}", 
                    exc_info=True)

def es_medico_productividad(legajo: Dict[str, Any]) -> bool:
    """
    Determina si es médico de productividad (Variables 1740, 1251, 1252).
    
    Condiciones acumulativas:
    - Puesto = 'MEDICO' (normalizado)
    - Sector principal está en lista de sectores médicos
    
    Args:
        legajo: Diccionario con datos del legajo
        
    Returns:
        bool: True si cumple criterios, False en caso contrario
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # 1. Obtener y normalizar puesto
        puesto_raw = legajo.get('datos_personales', {}).get('puesto')
        if puesto_raw is None:
            logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: ✗ NO APLICA - Puesto es None")
            return False
        
        puesto_normalizado = normalizar_texto(puesto_raw)
        logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: Puesto = '{puesto_raw}' (normalizado: '{puesto_normalizado}')")
        
        # 2. Verificar si puesto es MEDICO
        es_medico = puesto_normalizado == PUESTOS_ESPECIALES['MEDICO']
        logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: ¿Puesto == 'MEDICO'? {es_medico}")
        
        if not es_medico:
            logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: ✗ NO APLICA - Puesto no es MEDICO")
            return False
        
        # 3. Obtener y normalizar sector principal
        sector_raw = legajo.get('datos_personales', {}).get('sector', {}).get('principal')
        if sector_raw is None:
            logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: ✗ NO APLICA - Sector principal es None")
            return False
        
        sector_normalizado = normalizar_texto(sector_raw)
        logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: Sector = '{sector_raw}' (normalizado: '{sector_normalizado}')")
        
        # 4. Verificar si sector está en lista de sectores médicos
        en_sector_medico = sector_normalizado in SECTORES_MEDICOS
        logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: ¿Sector en SECTORES_MEDICOS? {en_sector_medico}")
        
        if en_sector_medico:
            logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: ✓ APLICA - Médico de productividad")
        else:
            logger.debug(f"[V1740/V1251/V1252] Legajo {id_legajo}: ✗ NO APLICA - Sector '{sector_raw}' no está en lista")
        
        return en_sector_medico
        
    except Exception as e:
        logger.error(f"[V1740/V1251/V1252] Legajo {id_legajo}: Error validando médico productividad - {str(e)}")
        return False

def es_licenciado_bioimagenes(legajo: Dict[str, Any]) -> bool:
    """
    Determina si aplica variable 10000 (Licenciado en Bioimágenes).
    
    Condiciones acumulativas:
    - Puesto en lista de puestos válidos (ConfigBioimagenes.PUESTOS_VALIDOS)
    - Sector en lista de sectores 156hs (SECTORES_ESPECIALES['HORAS_156'])
    - Adicionables contiene algún término de bioimágenes (ConfigBioimagenes.TERMINOS_ADICIONALES)

    Args:
        legajo: Diccionario con datos del legajo

    Returns:
        bool: True si cumple todas las condiciones, False en caso contrario
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    logger.debug(f"[V10000] Legajo {id_legajo}: Evaluando Licenciado en Bioimágenes")

    try:
        # 1. Obtener y normalizar puesto
        puesto_raw = legajo.get('datos_personales', {}).get('puesto')
        if puesto_raw is None:
            logger.debug(f"[V10000] Legajo {id_legajo}: ✗ NO APLICA - Puesto es None")
            return False
        
        puesto_normalizado = normalizar_texto(puesto_raw)
        logger.debug(f"[V10000] Legajo {id_legajo}: Puesto = '{puesto_raw}' (normalizado: '{puesto_normalizado}')")
        
        # 2. Verificar puesto en lista válida
        puesto_cumple = puesto_normalizado in ConfigBioimagenes.PUESTOS_VALIDOS
        logger.debug(f"[V10000] Legajo {id_legajo}: ¿Puesto en PUESTOS_VALIDOS? {puesto_cumple}")
        
        if not puesto_cumple:
            logger.debug(f"[V10000] Legajo {id_legajo}: ✗ NO APLICA - Puesto '{puesto_normalizado}' no válido")
            return False

        # 3. Obtener y normalizar sector principal
        sector_data = legajo.get('datos_personales', {}).get('sector')
        if sector_data is None or not isinstance(sector_data, dict):
            logger.debug(f"[V10000] Legajo {id_legajo}: ✗ NO APLICA - Datos sector inválidos")
            return False
        
        sector_principal_raw = sector_data.get('principal')
        if sector_principal_raw is None:
            logger.debug(f"[V10000] Legajo {id_legajo}: ✗ NO APLICA - Sector principal es None")
            return False
        
        sector_principal_normalizado = normalizar_texto(sector_principal_raw)
        logger.debug(f"[V10000] Legajo {id_legajo}: Sector = '{sector_principal_raw}' (normalizado: '{sector_principal_normalizado}')")

        # 4. Verificar sector en lista 156hs
        sector_cumple = sector_principal_normalizado in SECTORES_ESPECIALES.get('HORAS_156', [])
        logger.debug(f"[V10000] Legajo {id_legajo}: ¿Sector en HORAS_156? {sector_cumple}")
        
        if not sector_cumple:
            logger.debug(f"[V10000] Legajo {id_legajo}: ✗ NO APLICA - Sector '{sector_principal_normalizado}' no es 156hs")
            return False

        # 5. Obtener y normalizar adicionables
        adicionables_raw = legajo.get('remuneracion', {}).get('adicionables')
        adicionables_normalizado = normalizar_texto(adicionables_raw)
        logger.debug(f"[V10000] Legajo {id_legajo}: Adicionables = '{adicionables_raw}' (normalizado: '{adicionables_normalizado}')")

        # 6. Verificar términos en adicionables
        terminos_encontrados = [t for t in ConfigBioimagenes.TERMINOS_ADICIONALES if t in adicionables_normalizado]
        termino_adicional_cumple = len(terminos_encontrados) > 0
        logger.debug(f"[V10000] Legajo {id_legajo}: Términos encontrados: {terminos_encontrados}")
        logger.debug(f"[V10000] Legajo {id_legajo}: ¿Contiene término bioimágenes? {termino_adicional_cumple}")
        
        if not termino_adicional_cumple:
            logger.debug(f"[V10000] Legajo {id_legajo}: ✗ NO APLICA - Sin términos de bioimágenes en adicionables")
            return False

        # 7. Todas las condiciones cumplidas
        logger.info(f"[V10000] Legajo {id_legajo}: ✓ APLICA - Licenciado en Bioimágenes")
        return True

    except KeyError as ke:
        logger.error(f"[V10000] Legajo {id_legajo}: Error de clave (KeyError) - {str(ke)}")
        logger.error(traceback.format_exc())
        return False
    except Exception as e:
        logger.error(f"[V10000] Legajo {id_legajo}: Error inesperado - {str(e)}")
        logger.error(traceback.format_exc())
        return False

def calcular_horas_mensuales(legajo: Dict[str, Any], v239: float) -> float:
    """
    Calcula la variable 4 - Horas mensuales según reglas específicas.
    Aplica lógica robusta con normalización y control de errores.
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    try:
        # 1. Acceso seguro y normalización
        datos = legajo.get("datos_personales", {})
        puesto = normalizar_texto(datos.get("puesto")) # <--- Aquí se normaliza el 'puesto' del legajo
        sector = normalizar_texto(datos.get("sector", {}).get("principal"))

        logger.debug(f"[V4] Legajo {id_legajo}: INICIO EVALUACIÓN")
        logger.debug(f"[V4] Legajo {id_legajo}: ✓ Puesto raw='{datos.get('puesto')}' → normalizado='{puesto}'")
        logger.debug(f"[V4] Legajo {id_legajo}: ✓ Sector normalizado='{sector}'")
        logger.debug(f"[V4] Legajo {id_legajo}: ✓ v239 (horas semanales)={v239}")

        # 2. Casos especiales de 200 hs
        condicion_1 = (sector == "cuat" and puesto == PUESTOS_ESPECIALES['TELEFONISTA'] and v239 == 35)
        condicion_2 = (puesto == PUESTOS_ESPECIALES['RECEP_LAB'] and v239 == 35)
        condicion_3 = (puesto == PUESTOS_ESPECIALES['TEC_CARDIO'] and v239 >= 35)
        condicion_4 = (puesto == PUESTOS_ESPECIALES['OP_LOGISTICA'] and v239 >= 35)
        condicion_5 = (sector == "atencion al cliente laboratorio" and puesto == "recepcionista" and v239 >= 35)
        condicion_6 = (puesto == normalizar_texto("asistente tecnico") and v239 == 35)
        
        if condicion_1 or condicion_2 or condicion_3 or condicion_4 or condicion_5 or condicion_6:
            logger.debug(f"[V4] Legajo {id_legajo}: ✓ Cumple caso especial 200hs:")
            logger.debug(f"[V4] Legajo {id_legajo}:   - CUAT+Telefonista+35h: {condicion_1}")
            logger.debug(f"[V4] Legajo {id_legajo}:   - Recep Lab+35h: {condicion_2}")
            logger.debug(f"[V4] Legajo {id_legajo}:   - Téc Cardio+35h+: {condicion_3}")
            logger.debug(f"[V4] Legajo {id_legajo}:   - Op Logística+35h+: {condicion_4}")
            logger.debug(f"[V4] Legajo {id_legajo}:   - AtencLab+Recep+35h+: {condicion_5}")
            logger.debug(f"[V4] Legajo {id_legajo}:   - Asist Téc+35h: {condicion_6}")
            logger.info(f"[V4] Legajo {id_legajo}: ✓ RESULTADO = 200.00 horas")
            return 200.00
        else:
            logger.debug(f"[V4] Legajo {id_legajo}: ✗ No cumple casos especiales 200hs")

        # 3. Casos de puestos con piso 27 horas (bioquímicos, técnicos, etc.)
        puestos_piso_27 = [normalizar_texto(p) for p in [
            "AUXILIAR TECNICO", "TECNICO DE LABORATORIO",
            "TECNICO EXTRACCIONISTA", "BIOQUIMICO"
        ]]

        logger.debug(f"[V4] Legajo {id_legajo}: Evaluando puestos piso 27h")
        logger.debug(f"[V4] Legajo {id_legajo}:   - ¿Puesto en lista?: {puesto in puestos_piso_27}")
        
        if puesto in puestos_piso_27:
            logger.debug(f"[V4] Legajo {id_legajo}: ✓ Puesto con piso 27 reconocido: '{puesto}'")
            if 27 <= v239 <= 36:  # ✅ Rango exacto 27-36 → 156 horas
                logger.debug(f"[V4] Legajo {id_legajo}: ✓ v239={v239} está en rango [27-36]")
                logger.info(f"[V4] Legajo {id_legajo}: ✓ RESULTADO = 156.00 horas")
                return 156.00
            elif v239 < 27:  # ✅ Menos de 27 → proporcional 27 × 4.33
                horas_proporcionales = round(27 * 4.33, 2)
                logger.debug(f"[V4] Legajo {id_legajo}: ✓ v239={v239} < 27 → proporcional (27 × 4.33)")
                logger.info(f"[V4] Legajo {id_legajo}: ✓ RESULTADO = {horas_proporcionales} horas")
                return horas_proporcionales
            else:  # ✅ Más de 36 → continúa al siguiente caso
                logger.debug(f"[V4] Legajo {id_legajo}: ✓ v239={v239} > 36, continúa evaluación")
        else:
            logger.debug(f"[V4] Legajo {id_legajo}: ✗ No es puesto piso 27")

        # 4. Casos de puestos técnicos con piso 18 horas
        logger.debug(f"[V4] Legajo {id_legajo}: Evaluando técnicos piso 18h")
        es_tecnico_pivot = puesto in [normalizar_texto("TECNICO"), normalizar_texto("TECNICO PIVOT")]
        no_es_lab_excluido = sector != SECTOR_EXCLUIDO_LABORATORIO
        en_rango_18_36 = 18 <= v239 <= 36
        
        logger.debug(f"[V4] Legajo {id_legajo}:   - ¿Es TECNICO/TECNICO PIVOT?: {es_tecnico_pivot}")
        logger.debug(f"[V4] Legajo {id_legajo}:   - ¿Sector != '{SECTOR_EXCLUIDO_LABORATORIO}'?: {no_es_lab_excluido}")
        logger.debug(f"[V4] Legajo {id_legajo}:   - ¿v239 en [18-36]?: {en_rango_18_36}")
        
        if es_tecnico_pivot and no_es_lab_excluido and en_rango_18_36:
            logger.info(f"[V4] Legajo {id_legajo}: ✓ RESULTADO = 156.00 horas (técnico válido)")
            return 156.00
        else:
            logger.debug(f"[V4] Legajo {id_legajo}: ✗ No cumple caso técnicos 156hs")

        # 5. Caso médicos (pago proporcional directo)
        logger.debug(f"[V4] Legajo {id_legajo}: Evaluando profesionales de salud")
        logger.debug(f"[V4] Legajo {id_legajo}:   - Puesto '{puesto}' en lista profesionales: {puesto in valores_profesionales_para_comparacion}")
        
        if puesto in valores_profesionales_para_comparacion:
            resultado_proporcional = round(v239 * 4.33, 2)
            logger.debug(f"[V4] Legajo {id_legajo}: ✓ Profesional de salud → {v239} × 4.33 = {resultado_proporcional}")
            logger.info(f"[V4] Legajo {id_legajo}: ✓ RESULTADO = {resultado_proporcional} horas")
            return resultado_proporcional
        else:
            logger.debug(f"[V4] Legajo {id_legajo}: ✗ No es profesional de salud")

        # 6. Caso general con pisos (nuevo criterio) - CORREGIDO
        piso = PISOS_HORARIOS.get(normalizar_texto("GENERAL"), 36.0)
        sector_normalizado = normalizar_texto(sector)
        puesto_normalizado = normalizar_texto(puesto)

        logger.debug(f"[V4] Legajo {id_legajo}: Determinando piso horario (inicial={piso}h)")
        
        # Definir sectores y puestos de laboratorio
        puestos_lab_piso_27 = [normalizar_texto(p) for p in [
            "AUXILIAR TECNICO", "TECNICO DE LABORATORIO", 
            "TECNICO EXTRACCIONISTA", "BIOQUIMICO"
        ]]

        sectores_laboratorio = [
            normalizar_texto('LABORATORIO'),
            normalizar_texto('ATENCION AL CLIENTE LABORATORIO'),
            normalizar_texto('LABORATORIO CLINICO'),
            normalizar_texto('ANALISIS CLINICOS')
        ]

        # 6.1 Sector LABORATORIO con puesto específico → piso 27
        es_sector_lab = any(sector_normalizado == s for s in sectores_laboratorio)
        es_puesto_lab_27 = puesto_normalizado in puestos_lab_piso_27
        
        logger.debug(f"[V4] Legajo {id_legajo}:   - ¿Sector laboratorio?: {es_sector_lab}")
        logger.debug(f"[V4] Legajo {id_legajo}:   - ¿Puesto lab piso 27?: {es_puesto_lab_27}")
        
        if es_sector_lab and es_puesto_lab_27:
            piso = 27.0
            logger.debug(f"[V4] Legajo {id_legajo}: ✓ Sector lab + puesto específico → piso={piso}h")

        # 6.2 Sector IMÁGENES con puesto válido
        elif (
            sector_normalizado in SECTORES_IMAGENES
            and puesto_normalizado in ConfigBioimagenes.PUESTOS_VALIDOS
        ):
            piso = PISOS_HORARIOS.get(normalizar_texto("IMAGENES"), 18.0)
            logger.debug(f"[V4] Legajo {id_legajo}: ✓ Sector imágenes → piso={piso}h")

        logger.debug(f"[V4] Legajo {id_legajo}: Piso final determinado = {piso}h")

        # 7. Si está por debajo del piso → proporcional
        if v239 < piso:
            resultado_piso = round(piso * 4.33, 2)
            logger.debug(f"[V4] Legajo {id_legajo}: ✓ v239={v239} < piso={piso} → proporcional ({piso} × 4.33)")
            logger.info(f"[V4] Legajo {id_legajo}: ✓ RESULTADO = {resultado_piso} horas")
            return resultado_piso
        else:
            logger.debug(f"[V4] Legajo {id_legajo}: ✗ v239={v239} NO está debajo del piso {piso}h")

        # 8. Caso general por defecto
        logger.info(f"[V4] Legajo {id_legajo}: ✓ RESULTADO = 200.00 horas (caso general)")
        return 200.00

    except Exception as e:
        logger.error(f"[V4] Legajo {id_legajo}: ERROR CRÍTICO - {str(e)}")
        logger.error(traceback.format_exc())
        return 200.00

def calcular_jornada_reducida(legajo: Dict[str, Any], es_guardia: bool) -> Optional[float]:
    """
    Calcula la variable 1167 (% de jornada reducida) con detección robusta de puestos especiales.
    Versión mejorada con manejo más robusto de categorías FC/PFC y excepción Medicina Nuclear + Asistente Técnico.
    """
    try:
        # --- Extracción de datos ---
        id_legajo = legajo.get('id_legajo', 'N/A')
        datos_personales = legajo.get('datos_personales', {})
        puesto = normalizar_texto(datos_personales.get('puesto', ''))
        sector = normalizar_texto(datos_personales.get('sector', {}).get('principal', ''))
        total_horas = legajo.get('horario', {}).get('resumen', {}).get('total_horas_semanales', 0.0)
        categoria = legajo.get('contratacion', {}).get('categoria', '')

        logger.debug(f"[1167] Legajo {id_legajo}: Categoría raw: '{categoria}'")

        # --- Validación mejorada de categorías FC/PFC ---
        if isinstance(categoria, str) and categoria.lower().replace(' ', '_') in {'pfc', 'fc_pfc'}:
            logger.debug(f"[1167] Legajo {id_legajo}: Excluido por categoría FC/PFC: '{categoria}'")
            return None

        # --- Validación de condiciones de exclusión ---
        if es_guardia:
            logger.debug(f"[1167] Legajo {id_legajo}: Excluido (es guardia)")
            return None
        if not puesto:
            logger.warning(f"[1167] Legajo {id_legajo}: Puesto no definido")
            return None
        if not sector:
            logger.warning(f"[1167] Legajo {id_legajo}: Sector no definido")
            return None

        # --- Detección robusta de puestos especiales ---
        if es_puesto_especial(puesto) and total_horas == 35.0:
            logger.debug(f"[1167] Legajo {id_legajo}: Excluido (puesto especial '{puesto}' con 35h)")
            return None
        
        # --- Excepción Asistente Técnico con 35hs (entra en piso 36) ---
        if puesto == normalizar_texto("asistente tecnico") and total_horas == 35.0:
            logger.debug(f"[1167] Legajo {id_legajo}: Excluido (Asistente Técnico con 35h - entra en piso 36)")
            return None

        # --- Determinar piso horario ---
        dias_trabajo = set(legajo.get('horario', {}).get('resumen', {}).get('dias_trabajo', []))
        
        # Lógica para la regla especial de 18 horas
        if total_horas == 18.0 and dias_trabajo.issuperset(DIAS_ESPECIALES):
            piso = 45.0
            resultado = round((total_horas / piso) * 100, 4)
            logger.info(f"[1167] Legajo {id_legajo}: APLICA (regla especial 18h en L/M/V → {resultado}%)")
            return resultado
        
        # --- Asignación de piso horario según sector y puesto (con excepción) ---
        puestos_lab_piso_27 = [normalizar_texto(p) for p in [
            "AUXILIAR TECNICO", "TECNICO DE LABORATORIO", 
            "TECNICO EXTRACCIONISTA", "BIOQUIMICO"
        ]]
        
        # SECTORES RELACIONADOS CON LABORATORIO
        sectores_laboratorio = [
            normalizar_texto('LABORATORIO'),
            normalizar_texto('ATENCION AL CLIENTE LABORATORIO'),
            normalizar_texto('LABORATORIO CLINICO'),
            normalizar_texto('ANALISIS CLINICOS')
        ]
        
        logger.debug(f"[1167] Legajo {id_legajo}: DEBUG - Sector normalizado: '{sector}'")
        logger.debug(f"[1167] Legajo {id_legajo}: DEBUG - Puesto normalizado: '{puesto}'")
        logger.debug(f"[1167] Legajo {id_legajo}: DEBUG - Sectores laboratorio: {sectores_laboratorio}")
        logger.debug(f"[1167] Legajo {id_legajo}: DEBUG - ¿Sector relacionado con laboratorio? {any(sector == s for s in sectores_laboratorio)}")
        logger.debug(f"[1167] Legajo {id_legajo}: DEBUG - ¿Puesto en lista? {puesto in puestos_lab_piso_27}")

        # Si es sector RELACIONADO CON LABORATORIO y puesto específico → piso 27
        if any(sector == s for s in sectores_laboratorio) and puesto in puestos_lab_piso_27:
            piso = 27.0
            logger.debug(f"[1167] Legajo {id_legajo}: Sector laboratorio + puesto técnico '{puesto}' → piso 27h")

        # --- Excepción Medicina Nuclear + Asistente Técnico ---
        elif sector == normalizar_texto("medicina nuclear") and puesto == normalizar_texto("asistente tecnico"):
            piso = PISOS_HORARIOS.get(normalizar_texto('GENERAL'), 36.0)
            logger.debug(f"[1167] Legajo {id_legajo}: EXCEPCIÓN → Medicina Nuclear + Asist. Téc. → piso {piso}h (general)")

        elif sector in SECTORES_IMAGENES:
            piso = PISOS_HORARIOS.get(normalizar_texto('IMAGENES'), 36.0)
            logger.debug(f"[1167] Legajo {id_legajo}: Sector IMÁGENES → piso {piso}h")
        else:
            # TODOS los demás casos (incluyendo laboratorio sin puesto técnico) → piso general 36h
            piso = PISOS_HORARIOS.get(normalizar_texto('GENERAL'), 36.0)
            logger.debug(f"[1167] Legajo {id_legajo}: Sector '{sector}' + puesto '{puesto}' → piso GENERAL {piso}h")

        logger.debug(f"[1167] Legajo {id_legajo}: Piso determinado: {piso}h")

        # --- Cálculo final del porcentaje ---
        if total_horas < piso:
            resultado = round((total_horas / piso) * 100, 4)
            logger.info(f"[1167] Legajo {id_legajo}: APLICA ({total_horas}h < {piso}h → {resultado}%)")
            return resultado
            
        logger.debug(f"[1167] Legajo {id_legajo}: No aplica ({total_horas}h >= {piso}h)")
        return None

    except Exception as e:
        logger.error(f"[1167] Legajo {legajo.get('id_legajo', 'N/A')}: Error - {str(e)}")
        logger.error(traceback.format_exc())
        return None

def calcular_jornada_art19(legajo: Dict[str, Any], horas_semanales: float) -> Optional[int]:
    """
    Determina si aplica variable 1416 (Jornada Art. 19).
    
    Condiciones acumulativas:
    - Categoría contiene prefijo definido (ConfigArt19.CATEGORIA_PREFIX)
    - Puesto en lista de puestos válidos (ConfigArt19.PUESTOS_VALIDOS)
    - Sector coincide con sector válido (ConfigArt19.SECTOR_VALIDO, si está definido)
    - Horas semanales > ConfigArt19.HORAS_MIN

    Args:
        legajo: Diccionario con datos del legajo
        horas_semanales: Valor de variable 239 (horas semanales)

    Returns:
        int | None: 1 si cumple condiciones, None si no aplica
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # 0. Validaciones básicas
        if not legajo or not isinstance(horas_semanales, (int, float)):
            logger.debug(f"[V1416] Legajo {id_legajo}: ✗ Datos inválidos")
            return None

        # 1. Validar categoría
        categoria_raw = legajo.get('contratacion', {}).get('categoria', '')
        categoria = normalizar_texto(categoria_raw)
        categoria_prefix = normalizar_texto(ConfigArt19.CATEGORIA_PREFIX)
        
        if categoria_prefix not in categoria:
            logger.debug(f"[V1416] Legajo {id_legajo}: ✗ Categoría '{categoria_raw}' sin prefijo '{ConfigArt19.CATEGORIA_PREFIX}'")
            return None

        # 2. Validar puesto
        puesto_raw = legajo.get('datos_personales', {}).get('puesto', '')
        puesto = normalizar_texto(puesto_raw)
        
        if puesto not in ConfigArt19.PUESTOS_VALIDOS:
            logger.debug(f"[V1416] Legajo {id_legajo}: ✗ Puesto '{puesto_raw}' no válido")
            return None

        # 3. Validar sector (si está definido)
        if hasattr(ConfigArt19, 'SECTOR_VALIDO'):
            sector_raw = legajo.get('datos_personales', {}).get('sector', {}).get('principal', '')
            sector = normalizar_texto(sector_raw)
            
            if sector != ConfigArt19.SECTOR_VALIDO:
                logger.debug(f"[V1416] Legajo {id_legajo}: ✗ Sector '{sector_raw}' != '{ConfigArt19.SECTOR_VALIDO}'")
                return None

        # 4. Validar horas semanales
        if horas_semanales <= ConfigArt19.HORAS_MIN:
            logger.debug(f"[V1416] Legajo {id_legajo}: ✗ Horas {horas_semanales} <= {ConfigArt19.HORAS_MIN}")
            return None

        # 5. Todas las condiciones cumplidas
        return 1

    except Exception as e:
        logger.error(f"[V1416] Legajo {id_legajo}: Error - {str(e)}")
        return None

def calcular_porcentaje_art19(legajo: Dict[str, Any], v239: float) -> Optional[float]:
    """
    Calcula variable 1599 - % adicional por extensión horaria (Art. 19).

    Condiciones acumulativas:
    - Categoría contiene 'dentro de convenio' (CATEGORIA_ART19_PREFIX)
    - Puesto en lista PUESTOS_ART19
    - Sector principal == SECTOR_ART19
    - Horas semanales en rango (36, 48]
    
    Cálculo:
    - Si horas == 48 → 25%
    - Si 36 < horas < 48 → 25% * (horas / 48)

    Args:
        legajo: Diccionario con datos del legajo
        v239: Valor de variable 239 (horas semanales)

    Returns:
        float | None: Porcentaje calculado (4 decimales) o None si no aplica
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    logger.debug(f"[V1599] Legajo {id_legajo}: Evaluando porcentaje art.19. V239 = {v239}")

    try:
        # 1. Extraer y normalizar puesto
        puesto_raw = legajo.get('datos_personales', {}).get('puesto')
        if puesto_raw is None:
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ NO APLICA - Puesto es None")
            return None
        
        puesto = normalizar_texto(puesto_raw)
        logger.debug(f"[V1599] Legajo {id_legajo}: Puesto = '{puesto_raw}' (normalizado: '{puesto}')")

        # 2. Extraer categoría (sin normalizar, usar lower())
        categoria_raw = legajo.get('contratacion', {}).get('categoria')
        if categoria_raw is None:
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ NO APLICA - Categoría es None")
        
        categoria = categoria_raw.lower()

        # 3. Extraer y normalizar sector principal
        sector_data = legajo.get('datos_personales', {}).get('sector')
        if sector_data is None or not isinstance(sector_data, dict):
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ Datos sector inválidos")
            return None

        sector_principal_raw = sector_data.get('principal')
        if sector_principal_raw is None:
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ Sector principal None")
            return None
        
        sector_principal = normalizar_texto(sector_principal_raw)

        # 4. Validar categoría
        if CATEGORIA_ART19_PREFIX not in categoria:
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ Categoría '{categoria_raw}' sin '{CATEGORIA_ART19_PREFIX}'")
            return None

        # 5. Validar puesto
        if puesto not in PUESTOS_ART19:
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ Puesto '{puesto_raw}' no válido")
            return None

        # 6. Validar sector
        if sector_principal != SECTOR_ART19:
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ Sector '{sector_principal_raw}' != '{SECTOR_ART19}'")
            return None

        # 7. Validar rango de horas (36, 48]
        if not (HORAS_MIN_ART19 < v239 <= HORAS_MAX_ART19):
            logger.debug(f"[V1599] Legajo {id_legajo}: ✗ Horas {v239} fuera de rango ({HORAS_MIN_ART19}, {HORAS_MAX_ART19}]")
            return None

        # 8. Calcular porcentaje
        if v239 == HORAS_MAX_ART19:  # 48 horas exactas
            porcentaje = CONSTANTES['PORCENTAJE_MAX_ART19']
        else:  # Entre 36 y 48 (proporcional)
            porcentaje = CONSTANTES['PORCENTAJE_MAX_ART19'] * (v239 / HORAS_BASE_CALCULO_ART19)

        resultado = round(porcentaje, 4)
        return resultado

    except KeyError as ke:
        logger.error(f"[V1599] Legajo {id_legajo}: Falta campo - {str(ke)}")
        return None
    except TypeError as te:
        logger.error(f"[V1599] Legajo {id_legajo}: Error de tipo - {str(te)}")
        return None
    except Exception as e:
        logger.error(f"[V1599] Legajo {id_legajo}: Error - {str(e)}")
        return None


def calcular_extension_horaria(legajo: Dict[str, Any], v239: float) -> Optional[float]:
    """
    Calcula la extensión horaria (Variable 992) según reglas actualizadas:
    - La variable 992 DEBE SER IGUAL A LA VARIABLE 239 (horas semanales)
    - Aplica exclusivamente a:
      * Puestos: 'TÉCNICO' o 'TÉCNICO PIVOT'
      * Sectores: Deben estar en 'SECTORES_IMAGENES' y NO ser 'LABORATORIO'
      * Horas semanales > 24
      * Legajos con ID fuera del rango 4000-4999 (se excluye ese rango)

    Args:
        legajo: Diccionario con los datos completos del legajo
        v239: Valor ya calculado de la variable 239 (horas semanales)

    Returns:
        float: El mismo valor que v239 si cumple todas las condiciones
        None: Si no aplica el adicional

    Ejemplo:
        >>> # Asumiendo un legajo_ejemplo con datos válidos y v239=32.5
        >>> # calcular_extension_horaria(legajo_ejemplo, 32.5)
        # 32.5  # Para un técnico en mamografía con 32.5 horas semanales
    """
    id_legajo = legajo.get('id_legajo', 'DESCONOCIDO')
    logger.debug(f"Evaluando extensión horaria (992) para legajo ID: {id_legajo}")

    try:
        # =============================================
        # 1. VALIDACIONES INICIALES (con logging detallado y acceso seguro a datos)
        # =============================================

        # Validar ID de legajo
        if id_legajo == 'DESCONOCIDO' or not isinstance(id_legajo, int):
            logger.debug(f"Legajo {id_legajo} excluido (ID no válido)")
            return None
        if ConfigExtensionHoraria.ID_LEGAJO_EXCLUIDO_MIN <= id_legajo <= ConfigExtensionHoraria.ID_LEGAJO_EXCLUIDO_MAX:
            logger.debug(f"Legajo {id_legajo} excluido (ID en rango 4000-4999)")
            return None

        # Acceder y normalizar puesto de forma segura
        puesto_raw = legajo.get('datos_personales', {}).get('puesto')
        if puesto_raw is None:
            logger.debug(f"Legajo {id_legajo} excluido (puesto es None)")
            return None
        puesto_normalizado = normalizar_texto(puesto_raw)

        # Validar puesto (debe estar en los puestos válidos)
        if puesto_normalizado not in ConfigExtensionHoraria.PUESTOS_VALIDOS:
            logger.debug(f"Legajo {id_legajo} excluido (puesto '{puesto_normalizado}' no aplica para extensión horaria)")
            return None

        # Acceder y normalizar sector de forma segura
        sector_data = legajo.get('datos_personales', {}).get('sector', {})
        sector_principal_raw = sector_data.get('principal')
        if sector_principal_raw is None:
            logger.debug(f"Legajo {id_legajo} excluido (sector principal es None)")
            return None
        sector_normalizado = normalizar_texto(sector_principal_raw)

        # Validar sector: debe estar en SECTORES_IMAGENES y NO ser LABORATORIO
        if sector_normalizado not in SECTORES_IMAGENES:
            logger.debug(f"Legajo {id_legajo} excluido (sector '{sector_normalizado}' no está en SECTORES_IMAGENES)")
            return None

        if sector_normalizado == SECTOR_EXCLUIDO_LABORATORIO:
            logger.debug(f"Legajo {id_legajo} excluido (sector '{sector_normalizado}' es LABORATORIO)")
            return None

        # Validar horas mínimas
        if v239 <= 24:
            logger.debug(f"Legajo {id_legajo} excluido (horas semanales ({v239}) <= 24)")
            return None

        # =============================================
        # 2. APLICACIÓN DE REGLA PRINCIPAL
        # =============================================
        # REGLA CLAVE: 992 = 239 (mismo valor)
        valor_992 = round(float(v239), 2)

        logger.info(f"Legajo {id_legajo} CALCULA extensión horaria (992): {valor_992} (idéntico a 239 por regla)")

        return valor_992

    except KeyError as ke:
        logger.error(f"Legajo {id_legajo}: Falta campo obligatorio al calcular extensión horaria (992). Detalle: {str(ke)}")
        logger.error(traceback.format_exc())
        return None
    except Exception as e:
        logger.error(f"Legajo {id_legajo}: Error inesperado al calcular extensión horaria (992). Detalle: {str(e)}")
        logger.error(traceback.format_exc())
        return None

def calcular_adicional_pivot(legajo: Dict[str, Any]) -> Dict[int, int]:
    """
    Calcula el adicional pivot según puesto/sector.

    Reglas:
    - Puesto exacto: TECNICO PIVOT
    - Sector RESONANCIA MAGNETICA -> Variable 1145 = 40
    - Sectores parametrizados de imágenes -> Variable 1144 = 20
    """
    id_legajo = legajo.get('id_legajo', 'N/A')

    try:
        puesto_raw = legajo.get('datos_personales', {}).get('puesto')
        if puesto_raw is None:
            logger.debug(f"[V1145/V1144] Legajo {id_legajo}: Puesto es None")
            return {}

        puesto_normalizado = normalizar_texto(puesto_raw)
        if puesto_normalizado != ConfigAdicionalPivot.PUESTO_VALIDO:
            logger.debug(f"[V1145/V1144] Legajo {id_legajo}: Puesto '{puesto_normalizado}' no aplica")
            return {}

        sector_raw = legajo.get('datos_personales', {}).get('sector', {}).get('principal')
        if sector_raw is None:
            logger.debug(f"[V1145/V1144] Legajo {id_legajo}: Sector principal es None")
            return {}

        sector_normalizado = normalizar_texto(sector_raw)

        if sector_normalizado == ConfigAdicionalPivot.SECTOR_RESONANCIA:
            logger.info(f"[V1145] Legajo {id_legajo}: APLICA adicional pivot resonancia")
            return {ConfigAdicionalPivot.VARIABLE_RESONANCIA: ConfigAdicionalPivot.VALOR_RESONANCIA}

        if sector_normalizado in ConfigAdicionalPivot.SECTORES_VARIABLE_1144:
            logger.info(f"[V1144] Legajo {id_legajo}: APLICA adicional pivot general")
            return {ConfigAdicionalPivot.VARIABLE_GENERAL: ConfigAdicionalPivot.VALOR_GENERAL}

        logger.debug(f"[V1145/V1144] Legajo {id_legajo}: Sector '{sector_normalizado}' no aplica")
        return {}

    except Exception as e:
        logger.error(f"[V1145/V1144] Legajo {id_legajo}: Error calculando adicional pivot - {str(e)}")
        logger.error(traceback.format_exc())
        return {}

def calcular_adicional_resonancia(legajo: Dict[str, Any], v239: float) -> Optional[Any]:
    """
    Calcula la variable 1151 - Adicional Resonancia Magnética.
    
    Aplica si:
    - Puesto: TECNICO, TECNICO DE REPROCESO o TECNICO PIVOT
    - Sector: RESONANCIA MAGNETICA
    - Horas semanales coinciden con tabla de equivalencias
    
    Tabla de equivalencias:
    12hs -> 1, 24hs -> 2, 30hs -> 3, 34hs -> 4, 35hs -> 5
    36hs -> 6, 40hs -> 7, 45hs -> 8, 32.5hs -> 9
    
    Args:
        legajo: Diccionario con los datos completos del legajo
        v239: Valor de horas semanales (Variable 239)
    
    Returns:
        int: Valor según tabla de equivalencias
        str: Mensaje de error si las horas no coinciden con la tabla
        None: Si no aplica el adicional
    """
    id_legajo = legajo.get('id_legajo', 'N/A')
    
    try:
        # Tabla de equivalencias horas -> valor
        TABLA_RESONANCIA = {
            12.0: 1,
            24.0: 2,
            30.0: 3,
            34.0: 4,
            35.0: 5,
            36.0: 6,
            40.0: 7,
            45.0: 8,
            32.5: 9
        }
        
        # 1. Validar puesto
        puesto_raw = legajo.get('datos_personales', {}).get('puesto')
        if puesto_raw is None:
            logger.debug(f"[1151] Legajo {id_legajo}: Puesto es None")
            return None
        
        puesto_normalizado = normalizar_texto(puesto_raw)
        
        if puesto_normalizado not in ConfigBioimagenes.PUESTOS_VALIDOS:
            logger.debug(f"[1151] Legajo {id_legajo}: Puesto '{puesto_normalizado}' no aplica")
            return None
        
        # 2. Validar sector
        sector_data = legajo.get('datos_personales', {}).get('sector', {})
        sector_principal_raw = sector_data.get('principal')
        if sector_principal_raw is None:
            logger.debug(f"[1151] Legajo {id_legajo}: Sector principal es None")
            return None
        
        sector_normalizado = normalizar_texto(sector_principal_raw)
        
        if sector_normalizado != normalizar_texto("resonancia magnetica"):
            logger.debug(f"[1151] Legajo {id_legajo}: Sector '{sector_normalizado}' no es Resonancia Magnética")
            return None
        
        # 3. Buscar en tabla de equivalencias
        if v239 in TABLA_RESONANCIA:
            valor = TABLA_RESONANCIA[v239]
            logger.info(f"[1151] Legajo {id_legajo}: APLICA Adicional Resonancia - {v239}hs -> valor {valor}")
            return valor
        else:
            # No existe equivalencia para esas horas
            mensaje = f"No existe equivalencia de Adic Resonancia para esas hs semanales ({v239}hs)"
            logger.warning(f"[1151] Legajo {id_legajo}: {mensaje}")
            return mensaje
    
    except Exception as e:
        logger.error(f"[1151] Legajo {id_legajo}: Error calculando adicional resonancia - {str(e)}")
        logger.error(traceback.format_exc())
        return None

def calcular_dias_especiales(legajo: Dict[str, Any], v1242: int) -> Optional[int]:
    """
    Calcula variable 1131 - Días mensuales especiales.
    
    Condiciones evaluadas en orden:
    1. Horario Sadofe exacto ([5,6,7]) → retorna 10
    2. Horario Lu-Ma-Mi exacto ([0,1,2]) → retorna 10
    3. V1242 < 22 O puesto profesional O trabaja día 7 (feriado) → retorna v1242
    4. Ninguna condición cumplida → None
    
    Args:
        legajo: Diccionario con datos del legajo
        v1242: Valor de variable 1242 (días trabajados)
        
    Returns:
        int | None: 10, v1242, o None según condiciones
    """
    id_legajo = legajo.get('id_legajo', 'N/A')

    try:
        # 1. Obtener y normalizar datos
        datos = legajo.get("datos_personales", {})
        puesto = normalizar_texto(datos.get("puesto"))
        dias_semana_set = set(legajo.get("horario", {}).get("resumen", {}).get("dias_trabajo", []))

        # 2. Condición Especial: Horario Sadofe
        if dias_semana_set == {5, 6, 7}:
            return 10
            
        # 3. Condición Especial: Horario Lu-Ma-Mi
        if dias_semana_set == {0, 1, 2}:
            return 10

        # 4. Otras condiciones
        if v1242 < 22 or puesto in valores_profesionales_para_comparacion or 7 in dias_semana_set:
            return v1242

        # 5. No aplica
        logger.debug(f"[V1131] Legajo {id_legajo}: ✗ Días={dias_semana_set}, V1242={v1242}")
        return None
        
    except Exception as e:
        logger.error(f"[V1131] Legajo {id_legajo}: Error - {str(e)}")
        return None

def aplicar_proporcion_lavado(legajo: Dict[str, Any]) -> bool:
    """
    Determina si aplica el adicional de lavado de uniforme (Variable 1673).

    Condiciones acumulativas:
    1. Puesto = "Operario de Logística" (normalizado)
    2. Subsector = "Interior" (normalizado)
    3. Total horas semanales < 35

    Args:
        legajo: Diccionario completo del legajo

    Returns:
        bool: True si aplica, False en caso contrario
    """
    id_legajo = legajo.get('id_legajo', 'N/A')

    try:
        # 1. Validar y extraer datos
        datos_personales = legajo.get('datos_personales')
        if not isinstance(datos_personales, dict):
            logger.debug(f"[V1673] Legajo {id_legajo}: ✗ datos_personales inválido")
            return False

        puesto_normalizado = normalizar_texto(datos_personales.get('puesto'))
        
        if puesto_normalizado != normalizar_texto("OPERARIO DE LOGISTICA"):
            logger.debug(f"[V1673] Legajo {id_legajo}: ✗ Puesto no es 'Operario de Logística'")
            return False

        sector_data = datos_personales.get('sector')
        if not isinstance(sector_data, dict):
            logger.debug(f"[V1673] Legajo {id_legajo}: ✗ sector inválido")
            return False

        subsector_normalizado = normalizar_texto(sector_data.get('subsector'))
        
        if subsector_normalizado != normalizar_texto("INTERIOR"):
            logger.debug(f"[V1673] Legajo {id_legajo}: ✗ Subsector no es 'Interior'")
            return False

        # 2. Validar horas
        horas_raw = legajo.get('horario', {}).get('resumen', {}).get('total_horas_semanales')
        
        if horas_raw is None:
            logger.debug(f"[V1673] Legajo {id_legajo}: ✗ total_horas_semanales None")
            return False

        try:
            total_horas = float(horas_raw)
        except (ValueError, TypeError):
            logger.debug(f"[V1673] Legajo {id_legajo}: ✗ Horas inválidas")
            return False

        if total_horas >= 35.0:
            logger.debug(f"[V1673] Legajo {id_legajo}: ✗ Horas {total_horas} >= 35")
            return False

        return True

    except Exception as e:
        logger.error(f"[V1673] Legajo {id_legajo}: Error - {e}")
        return False

# ==============================
# FUNCIONES DE REPORTE Y SALIDA
# ==============================

def generar_reporte_parcial(
    estadisticas: Dict[str, Any],
    ruta_archivo_procesado: Optional[str] = None
) -> None:
    """
    Genera un reporte parcial de procesamiento ultra-perfeccionado
    con formato, color (para terminales compatibles) y detalles adicionales.

    Args:
        estadisticas: Diccionario con las métricas de procesamiento.
                      Se espera que contenga al menos:
                      'total_legajos', 'legajos_procesados',
                      'legajos_con_error', 'variables_calculadas'.
                      Los valores faltantes serán tratados como 0.
        ruta_archivo_procesado: Ruta opcional del archivo JSON/origen que fue procesado.
                                Si se proporciona, se incluirá en el reporte.
    """
    try:
        # Acceso robusto a las estadísticas usando .get() con valores por defecto.
        total_legajos = estadisticas.get('total_legajos', 0)
        legajos_procesados = estadisticas.get('legajos_procesados', 0)
        legajos_con_error = estadisticas.get('legajos_con_error', 0)
        variables_calculadas = estadisticas.get('variables_calculadas', 0)

        # --- Cálculo de la Tasa de Éxito ---
        tasa_exito_str = "0.00%"
        tasa_exito_color = COLOR_GREEN
        if total_legajos > 0:
            try:
                tasa_exito = (legajos_procesados / total_legajos) * 100
                tasa_exito_str = f"{tasa_exito:.2f}%"

                if tasa_exito == 100:
                    tasa_exito_color = COLOR_GREEN
                elif tasa_exito >= 80:
                    tasa_exito_color = COLOR_YELLOW
                else:
                    tasa_exito_color = COLOR_RED
            except Exception as e:
                logger.error(f"Error inesperado al calcular la tasa de éxito: {e}", exc_info=True)
                tasa_exito_str = "Error cálculo"
                tasa_exito_color = COLOR_RED
        else:
            tasa_exito_color = COLOR_YELLOW

        # --- Determinación del Estado General del Procesamiento ---
        estado_general_mensaje = ""
        estado_general_color = COLOR_RESET
        if total_legajos == 0:
            estado_general_mensaje = "NO SE ENCONTRARON DATOS PARA PROCESAR"
            estado_general_color = COLOR_YELLOW
        elif legajos_con_error > 0 and legajos_procesados == 0:
            estado_general_mensaje = "FALLO CRÍTICO: NINGÚN LEGAJO PROCESADO CORRECTAMENTE"
            estado_general_color = COLOR_RED
        elif legajos_con_error > 0:
            estado_general_mensaje = "PROCESAMIENTO COMPLETADO CON ERRORES DETECTADOS"
            estado_general_color = COLOR_YELLOW
        else:
            estado_general_mensaje = "PROCESAMIENTO COMPLETO Y EXITOSO"
            estado_general_color = COLOR_GREEN

        # --- Construcción del Reporte Final con Formato y Colores ---
        reporte = f"""
{COLOR_BOLD}{COLOR_CYAN}╔═══════════════════════════════════════════════════════════╗{COLOR_RESET}
{COLOR_BOLD}{COLOR_CYAN}║         INFORME PARCIAL DE PROCESAMIENTO DE LEGAJOS       ║{COLOR_RESET}
{COLOR_BOLD}{COLOR_CYAN}╚═══════════════════════════════════════════════════════════╝{COLOR_RESET}
{COLOR_BLUE}Fecha del Reporte:{COLOR_RESET} {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
{COLOR_BLUE}Archivo Procesado:{COLOR_RESET} {ruta_archivo_procesado if ruta_archivo_procesado else 'N/A (No especificado)'}
{COLOR_BOLD}{COLOR_CYAN}─────────────────────────────────────────────────────────────{COLOR_RESET}

{COLOR_BOLD}≫ ESTADÍSTICAS CLAVE:{COLOR_RESET}
  • Total de legajos a procesar:   {total_legajos}
  • Legajos procesados exitosamente: {COLOR_GREEN}{legajos_procesados}{COLOR_RESET}
  • Legajos con errores detectados:  {COLOR_RED}{legajos_con_error}{COLOR_RESET}
  • Variables calculadas generadas:  {COLOR_BLUE}{variables_calculadas}{COLOR_RESET}

{COLOR_BOLD}≫ RENDIMIENTO GENERAL:{COLOR_RESET}
  • Tasa de éxito del procesamiento: {tasa_exito_color}{COLOR_BOLD}{tasa_exito_str}{COLOR_RESET}

{COLOR_BOLD}{COLOR_CYAN}─────────────────────────────────────────────────────────────{COLOR_RESET}
{COLOR_BOLD}≫ ESTADO DEL PROCESAMIENTO:{COLOR_RESET} {estado_general_color}{COLOR_BOLD}{estado_general_mensaje}{COLOR_RESET}
{COLOR_BOLD}{COLOR_CYAN}─────────────────────────────────────────────────────────────{COLOR_RESET}

{COLOR_BLUE}Notas:{COLOR_RESET}
  - Para detalles de errores, revise el archivo 'liquidacion_debug.log'.
  - Los archivos de resultados CSV contienen las variables generadas.
"""
        logger.info(reporte)
        print(reporte)

    except Exception as e:
        logger.error(f"Error CRÍTICO al generar el reporte parcial. Detalle: {e}", exc_info=True)

def generar_reporte_final(resultados: List[Tuple[int, int, Any]], ruta_archivo: str) -> None:
    """Genera un reporte final detallado"""
    try:
        # Estadísticas por variable
        variables_calculadas = len(resultados)
        variables_unicas = len({v[1] for v in resultados})

        # Conteo por tipo de variable
        conteo_variables = {}
        for _, codigo, _ in resultados:
            conteo_variables[codigo] = conteo_variables.get(codigo, 0) + 1

        # Top 5 variables más frecuentes
        top_variables = sorted(conteo_variables.items(), key=lambda x: x[1], reverse=True)[:5]

        reporte = f"""
        INFORME FINAL DE PROCESAMIENTO
        ==============================
        Archivo procesado: {ruta_archivo}
        Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}

        ESTADÍSTICAS GENERALES
        ---------------------
        - Total variables calculadas: {variables_calculadas}
        - Variables únicas calculadas: {variables_unicas}

        VARIABLES MÁS FRECUENTES
        ------------------------
        {chr(10).join(f'- Variable {codigo}: {cantidad} veces' for codigo, cantidad in top_variables)}

        ARCHIVOS GENERADOS
        ------------------
        - variables_calculadas.csv: Contiene todas las variables calculadas
        - liquidacion_debug.log: Registro detallado del procesamiento

        REVISIONES RECOMENDADAS
        -----------------------
        1. Verificar legajos con errores en el log
        2. Validar variables con conteo inusual
        3. Revisar casos especiales (guardias, médicos, etc.)
        """
        logger.info(reporte)
        print(reporte)
        # Guardar reporte en archivo
        with open('reporte_final.txt', 'w', encoding='utf-8') as f:
            f.write(reporte)

    except Exception as e:
        logger.error(f"Error generando reporte final: {str(e)}")
        
# =============== BLOQUE DE EJECUCIÓN INDEPENDIENTE ===============
# =============== BLOQUE DE EJECUCIÓN INDEPENDIENTE ===============
if __name__ == '__main__':
    # Esta sección SÓLO se ejecuta cuando corres este archivo directamente.
    
    # 1. Se ha eliminado la configuración local de logging de aquí.
    #    Ahora, la configuración debe hacerse a nivel de la aplicación
    #    que use este script, como una app de Streamlit, para evitar
    #    conflictos y duplicación.
    #
    #    Ejemplo de configuración para una aplicación que importe este script:
    #    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')

    logger.info("--- Ejecutando json_a_excel.py en modo de prueba ---")
    
    # 2. Tu código de prueba
    try:
        # Crea un archivo JSON de prueba si no existe
        json_prueba = {
            "legajos": [
                {
                    "id_legajo": 101,
                    "remuneracion": {"sueldo_base": 1000},
                    "horario": {
                        "resumen": {"total_horas_semanales": 40, "total_horas_nocturnas": 0, "dias_trabajo": [0,1,2,3,4]},
                        "bloques": [{"dias_semana": [0,1,2,3,4], "hora_inicio": "09:00", "hora_fin": "17:00"}]
                    },
                    "contratacion": {"categoria": "dc_1_categoria"},
                    "datos_personales": {"sede": "Pilar", "sector": {"principal": "Administración"}}
                }
            ]
        }
        with open("horarios_prueba.json", "w") as f:
            json.dump(json_prueba, f)

        # Llama a tus funciones principales
        resultados, stats = procesar_archivo_json("horarios_prueba.json")
        if resultados:
            guardar_resultados_csv(resultados, "resultados_de_prueba.xlsx")
        
        generar_reporte_parcial(stats, "horarios_prueba.json")

    except Exception as e:
        logger.critical(f"Ocurrió un error catastrófico durante la prueba: {e}", exc_info=True)